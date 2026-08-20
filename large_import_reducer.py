#!/usr/bin/env python3
"""Streaming preprocessor for opt-in large cloud and on-prem imports.

The normal parsers intentionally remain unchanged. Cloud bills spill exact
aggregation state to SQLite; on-prem inventories are projected without merging
rows. Both paths emit normalized CSV that the existing parsers can consume.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
import tempfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, getcontext
from pathlib import Path


SOURCE_IMPORT_MODES = {
    "azure": "large_cloud_bill_reduced",
    "aws": "large_cloud_bill_reduced",
    "on_prem": "large_onprem_streamed",
}
AUDIT_COLUMNS = (
    "first_usage_date",
    "last_usage_date",
    "source_row_count",
    "source_file_name",
    "source_import_mode",
    "resource_identity_status",
)
AWS_COMPACT_AUDIT_COLUMNS = ("resource_identity_status",)
AWS_AGGREGATION_AUDIT_COLUMNS = (
    "first_usage_date",
    "last_usage_date",
    "source_row_count",
    "resource_identity_status",
)


def _norm(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


# Canonical names are used only for validation and reducer behavior.  The output
# retains the source header spelling so the existing cloud-bill aliases still see
# precisely the vocabulary they already support.
AZURE_ALIASES = {
    "Date": ("Date", "UsageDate", "Usage Date"),
    "BillingAccountName": ("BillingAccountName", "Billing Account Name"),
    "AccountName": ("AccountName", "Account Name"),
    "SubscriptionName": ("SubscriptionName", "Subscription Name"),
    "Product": ("Product", "ProductName", "Product Name"),
    "PartNumber": ("PartNumber", "Part Number"),
    "ServiceFamily": ("ServiceFamily", "Service Family"),
    "MeterCategory": ("MeterCategory", "Meter Category"),
    "MeterSubCategory": ("MeterSubCategory", "Meter SubCategory", "Meter Sub-Category"),
    "MeterRegion": ("MeterRegion", "Meter Region"),
    "MeterName": ("MeterName", "Meter Name"),
    "Quantity": ("Quantity", "ConsumedQuantity", "Consumed Quantity"),
    "EffectivePrice": ("EffectivePrice", "Effective Price"),
    "Cost": ("Cost", "CostInBillingCurrency", "Cost In Billing Currency", "PreTaxCost"),
    "UnitPrice": ("UnitPrice", "Unit Price"),
    "BillingCurrency": ("BillingCurrency", "BillingCurrencyCode", "Billing Currency", "Currency"),
    "ResourceLocation": ("ResourceLocation", "Resource Location", "Location"),
    "UnitOfMeasure": ("UnitOfMeasure", "Unit Of Measure", "Unit"),
    "PublisherName": ("PublisherName", "Publisher Name"),
    "PlanName": ("PlanName", "Plan Name"),
    "ChargeType": ("ChargeType", "Charge Type"),
    "Frequency": ("Frequency",),
    "PublisherType": ("PublisherType", "Publisher Type"),
    "PayGPrice": ("PayGPrice", "Pay G Price", "PAYG Price"),
    "PricingModel": ("PricingModel", "Pricing Model"),
    "ConsumedService": ("ConsumedService", "Consumed Service"),
    "AdditionalInfo": ("AdditionalInfo", "Additional Info"),
    "SubscriptionId": ("SubscriptionId", "Subscription ID"),
    "ResourceId": ("ResourceId", "Resource ID", "InstanceId", "Instance ID"),
    "ResourceGroup": ("ResourceGroup", "Resource Group", "ResourceGroupName"),
    "ResourceName": ("ResourceName", "Resource Name", "InstanceName", "Instance Name"),
    "Environment": ("Environment", "EnvironmentName", "Environment Name", "EnvironmentTag"),
    "DatabaseServerId": (
        "DatabaseServerId", "Database Server ID", "DatabaseId", "Database ID",
        "ServerId", "Server ID", "ServerName", "Server Name",
    ),
}

AWS_ALIASES = {
    "Date": (
        "lineItem/UsageStartDate", "lineItem UsageStartDate", "UsageStartDate",
        "bill/BillingPeriodStartDate", "BillingPeriodStartDate", "Date",
    ),
    "BillingAccountName": (
        "bill/PayerAccountId", "PayerAccountId", "Payer Account ID", "BillingAccountId",
    ),
    "AccountName": (
        "lineItem/UsageAccountId", "UsageAccountId", "LinkedAccountId", "AccountId",
    ),
    "SubscriptionName": ("LinkedAccountName", "AccountName"),
    "Product": (
        "product/ProductName", "ProductName", "product/ServiceName", "ServiceName",
        "lineItem/ProductCode", "ProductCode",
    ),
    "PartNumber": ("product/sku", "ProductSku", "Sku", "SKU"),
    "ServiceFamily": (
        "product/ProductFamily", "ProductFamily", "lineItem/ProductCode", "ProductCode",
    ),
    "MeterCategory": ("lineItem/ProductCode", "ProductCode", "ServiceCode"),
    "MeterSubCategory": ("lineItem/UsageType", "UsageType", "Operation"),
    "MeterRegion": ("product/region", "Region", "AvailabilityZone"),
    "MeterName": (
        "lineItem/LineItemDescription", "LineItemDescription", "lineItem/UsageType",
        "UsageType", "ItemDescription",
    ),
    "Quantity": ("lineItem/UsageAmount", "UsageAmount", "Usage Quantity", "Quantity"),
    "EffectivePrice": (
        "lineItem/NetUnblendedRate", "NetUnblendedRate", "lineItem/UnblendedRate",
        "UnblendedRate", "Rate",
    ),
    "Cost": (
        "lineItem/NetUnblendedCost", "NetUnblendedCost", "lineItem/UnblendedCost",
        "UnblendedCost", "lineItem/BlendedCost", "BlendedCost", "Cost",
        "CostBeforeTax", "TotalCost",
    ),
    "UnitPrice": ("pricing/publicOnDemandRate", "PublicOnDemandRate", "UnitPrice"),
    "BillingCurrency": ("pricing/currency", "PricingCurrency", "Currency", "CurrencyCode"),
    "ResourceLocation": (
        "product/region", "Region", "lineItem/AvailabilityZone", "AvailabilityZone",
    ),
    "UnitOfMeasure": ("pricing/unit", "PricingUnit", "UsageUnit", "Unit"),
    "ChargeType": ("lineItem/LineItemType", "LineItemType", "ChargeType"),
    "PricingModel": ("pricing/term", "PricingTerm", "PurchaseOption", "PricingModel"),
    "ResourceId": ("lineItem/ResourceId", "ResourceId", "Resource ID"),
    "ResourceName": ("ResourceName", "InstanceName"),
    "Environment": ("Environment", "EnvironmentTag"),
}

# These fields define the minimum safe Azure enrollment contract.  Optional
# pricing and publisher columns are retained whenever present but do not block an
# otherwise valid export.
AZURE_REQUIRED = (
    "Date",
    "BillingAccountName",
    "AccountName",
    "SubscriptionName",
    "Product",
    "PartNumber",
    "MeterCategory",
    "MeterName",
    "Quantity",
    "Cost",
    "BillingCurrency",
    "ResourceLocation",
    "UnitOfMeasure",
)

AWS_REQUIRED = (
    "Date",
    "BillingAccountName",
    "AccountName",
    "Product",
    "MeterName",
    "Quantity",
    "Cost",
    "BillingCurrency",
)

IDENTITY_FIELDS = ("ResourceId", "ResourceGroup", "ResourceName", "Environment", "DatabaseServerId")
IDENTITY_HEADER_TOKENS = (
    "resourceid", "resourceguid", "resourceuri", "resourcegroup", "resourcename",
    "databaseid", "databasename", "serverid", "servername", "instanceid",
    "instancename", "environment", "envtag",
)


class LargeImportReductionError(ValueError):
    pass


def _header_map(headers, aliases=None):
    aliases = aliases or AZURE_ALIASES
    normalized = {}
    for idx, header in enumerate(headers):
        normalized.setdefault(_norm(header), idx)
    resolved = {}
    for canonical, choices in aliases.items():
        for alias in choices:
            if _norm(alias) in normalized:
                resolved[canonical] = normalized[_norm(alias)]
                break
    return resolved


def _header_score(row, aliases=None, required=None):
    required = required or AZURE_REQUIRED
    resolved = _header_map(row, aliases)
    return sum(3 if name in required else 1 for name in resolved)


def _cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value, label, source_row):
    text = _cell(value).replace(",", "")
    if not text:
        return Decimal("0")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    text = text.replace("$", "").replace("€", "").replace("£", "").strip()
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise LargeImportReductionError(
            f"Row {source_row:,} has a non-numeric {label} value: {value!r}."
        ) from exc
    if not number.is_finite():
        raise LargeImportReductionError(
            f"Row {source_row:,} has an invalid {label} value: {value!r}."
        )
    return number


def _decimal_text(value):
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _decimal_add(left, right):
    return _decimal_text(Decimal(left or "0") + Decimal(right or "0"))


def _date_key(value):
    text = _cell(value)
    if not text:
        return "", ""
    iso_candidate = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(iso_candidate).isoformat(), text
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(text, fmt).isoformat(), text
        except ValueError:
            continue
    # Unknown formats remain deterministic. They are not interpreted or changed.
    return f"text:{text}", text


def _csv_preflight(path, scorer):
    with Path(path).open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        sample = fh.read(65536)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    with Path(path).open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        reader = csv.reader(fh, dialect)
        candidates = []
        for idx, row in enumerate(reader):
            candidates.append((idx, [_cell(v) for v in row]))
            if idx >= 24:
                break
    if not candidates:
        raise LargeImportReductionError("The file is empty.")
    header_index, headers = max(candidates, key=lambda item: scorer(item[1]))
    return dialect, header_index, headers, "Streamed CSV"


def _xlsx_preflight(path, sheet_name, scorer):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise LargeImportReductionError("openpyxl is required for streaming XLSX reduction.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise LargeImportReductionError(f"Worksheet {sheet_name!r} was not found.")
            worksheets = [workbook[sheet_name]]
        else:
            worksheets = [ws for ws in workbook.worksheets if ws.sheet_state == "visible"] or list(workbook.worksheets)
        choices = []
        for ws in worksheets:
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                values = [_cell(v) for v in row]
                choices.append((scorer(values), ws.title, idx, values))
                if idx >= 24:
                    break
        if not choices:
            raise LargeImportReductionError("The workbook has no readable rows.")
        _score, chosen_sheet, header_index, headers = max(choices, key=lambda item: item[0])
        return header_index, headers, chosen_sheet
    finally:
        workbook.close()


INVENTORY_IDENTITY_TOKENS = (
    "machine", "server", "hostname", "host", "vmname", "virtualmachine",
    "instance", "asset", "application", "workload", "database", "cluster",
    "environmentname", "enviornmentname",
)
INVENTORY_CPU_TOKENS = ("cpu", "vcpu", "ocpu", "core", "processorcount")
INVENTORY_MEMORY_TOKENS = ("memory", "ram", "memgb")
INVENTORY_RETAIN_TOKENS = (
    *INVENTORY_IDENTITY_TOKENS,
    *INVENTORY_CPU_TOKENS,
    *INVENTORY_MEMORY_TOKENS,
    "storage", "disk", "capacity", "allocated", "iops", "throughput",
    "environment", "env", "operatingsystem", "osname", "platform", "linux",
    "windows", "region", "location", "site", "datacenter", "availabilityzone",
    "resource", "subscription", "account", "owner", "department", "costcenter",
    "tier", "role", "engine", "edition", "version", "quantity", "count", "numberof",
    "network", "bandwidth", "gpu", "shape", "instance type", "instancetype",
)


def _inventory_header_signals(headers):
    normalized = [_norm(header) for header in headers]
    identity = [idx for idx, text in enumerate(normalized)
                if any(token in text for token in INVENTORY_IDENTITY_TOKENS)]
    cpu = [idx for idx, text in enumerate(normalized)
           if any(token in text for token in INVENTORY_CPU_TOKENS)
           and not any(bad in text for bad in ("utilization", "utilisation", "ratio", "rate", "percent"))]
    memory = [idx for idx, text in enumerate(normalized)
              if any(token in text for token in INVENTORY_MEMORY_TOKENS)
              and "storage" not in text and "utilization" not in text]
    retained = [idx for idx, text in enumerate(normalized)
                if any(token.replace(" ", "") in text for token in INVENTORY_RETAIN_TOKENS)]
    return {"identity": identity, "cpu": cpu, "memory": memory, "retained": retained}


def _inventory_header_score(headers):
    signals = _inventory_header_signals(headers)
    return (
        len(signals["retained"])
        + (20 if signals["identity"] else 0)
        + (20 if signals["cpu"] else 0)
        + (20 if signals["memory"] else 0)
    )


def preflight_large_import(path, import_kind, sheet_name=""):
    """Inspect only the header area and return a validated streaming plan."""
    import_kind = str(import_kind or "").strip().lower()
    if import_kind not in SOURCE_IMPORT_MODES:
        raise LargeImportReductionError(
            "Large import type must be on_prem, aws, or azure."
        )
    if import_kind == "azure":
        aliases, required = AZURE_ALIASES, AZURE_REQUIRED
        scorer = lambda row: _header_score(row, aliases, required)
    elif import_kind == "aws":
        aliases, required = AWS_ALIASES, AWS_REQUIRED
        scorer = lambda row: _header_score(row, aliases, required)
    else:
        aliases, required = None, None
        scorer = _inventory_header_score

    source = Path(path)
    suffix = source.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        dialect, header_index, headers, selected_sheet = _csv_preflight(source, scorer)
        reader_kind = "csv"
    elif suffix == ".xlsx":
        header_index, headers, selected_sheet = _xlsx_preflight(source, sheet_name, scorer)
        dialect = None
        reader_kind = "xlsx"
    else:
        raise LargeImportReductionError(
            "Large-file preprocessing accepts .xlsx, .csv, or .tsv exports only."
        )

    if import_kind == "on_prem":
        signals = _inventory_header_signals(headers)
        missing = [label for label, key in (
            ("machine/server/application identity", "identity"),
            ("CPU/core count", "cpu"),
            ("memory/RAM", "memory"),
        ) if not signals[key]]
        if missing:
            raise LargeImportReductionError(
                "This is not a supported on-prem inventory table. Missing: "
                + ", ".join(missing) + "."
            )
        resolved = {}
        retained_indices = signals["retained"]
        identity_indices = signals["identity"]
    else:
        resolved = _header_map(headers, aliases)
        missing = [name for name in required if name not in resolved]
        if missing:
            label = "Azure enrollment" if import_kind == "azure" else "AWS billing"
            raise LargeImportReductionError(
                f"This is not a supported {label} export. Missing required columns: "
                + ", ".join(missing) + "."
            )
        # Bill schemas vary, and an unfamiliar column can still be an audit or
        # identity boundary. Retain every source column and put every non-date,
        # non-summed value into the exact aggregation key.
        retained_indices = list(range(len(headers)))
        identity_indices = []

    return {
        "kind": reader_kind,
        "importKind": import_kind,
        "dialect": dialect,
        "headerIndex": header_index,
        "headers": headers,
        "sheetName": selected_sheet,
        "resolved": resolved,
        "retainedIndices": retained_indices,
        "identityIndices": identity_indices,
    }


def _iter_source_rows(path, plan):
    if plan["kind"] == "csv":
        with Path(path).open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
            reader = csv.reader(fh, plan["dialect"])
            for idx, row in enumerate(reader):
                if idx <= plan["headerIndex"]:
                    continue
                yield idx + 1, [_cell(v) for v in row]
        return

    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[plan["sheetName"]]
        for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if idx <= plan["headerIndex"]:
                continue
            yield idx + 1, [_cell(v) for v in row]
    finally:
        workbook.close()


def _write_audit(output, audit):
    audit_path = output.with_suffix(output.suffix + ".audit.json")
    audit["auditFilePath"] = str(audit_path)
    try:
        audit_path.write_text(
            json.dumps(audit, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    except OSError as exc:
        output.unlink(missing_ok=True)
        audit_path.unlink(missing_ok=True)
        raise LargeImportReductionError(
            f"Could not persist reducer audit metadata: {exc}"
        ) from exc
    return audit


def _reduce_large_bill(path, output_path, import_kind, sheet_name="", progress=None):
    """Stream an AWS or Azure bill into an exactly aggregated CSV."""
    source = Path(path)
    output = Path(output_path)
    plan = preflight_large_import(source, import_kind, sheet_name)
    source_import_mode = SOURCE_IMPORT_MODES[import_kind]
    headers = plan["headers"]
    resolved = plan["resolved"]
    retained_indices = plan["retainedIndices"]
    retained_headers = [headers[idx] for idx in retained_indices]
    date_idx = resolved["Date"]
    quantity_idx = resolved["Quantity"]
    cost_idx = resolved["Cost"]
    identity_indices = {resolved[name] for name in IDENTITY_FIELDS if name in resolved}
    identity_indices.update(
        idx for idx, header in enumerate(headers)
        if any(token in _norm(header) for token in IDENTITY_HEADER_TOKENS)
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    temp = tempfile.NamedTemporaryFile(prefix="cloud-bill-reducer-", suffix=".sqlite3", dir=output.parent, delete=False)
    temp.close()
    database_path = Path(temp.name)
    connection = sqlite3.connect(database_path)
    connection.create_function("decimal_add", 2, _decimal_add, deterministic=True)
    connection.executescript(
        """
        PRAGMA journal_mode=OFF;
        PRAGMA synchronous=OFF;
        PRAGMA temp_store=FILE;
        CREATE TABLE aggregates (
            key_hash TEXT PRIMARY KEY,
            key_json TEXT NOT NULL,
            row_json TEXT NOT NULL,
            quantity TEXT NOT NULL,
            cost TEXT NOT NULL,
            first_date_key TEXT NOT NULL,
            first_date TEXT NOT NULL,
            last_date_key TEXT NOT NULL,
            last_date TEXT NOT NULL,
            source_row_count INTEGER NOT NULL,
            identity_status TEXT NOT NULL,
            first_source_row INTEGER NOT NULL
        );
        """
    )
    decimal_context = getcontext()
    original_decimal_precision = decimal_context.prec
    # Excel values commonly arrive as 15-17 digit floats. Summing hundreds of
    # thousands in source order and then re-summing grouped results must remain
    # associative for the final reconciliation check; Decimal's default 28 digits
    # is not enough for the largest enrollment quantities.
    decimal_context.prec = max(original_decimal_precision, 80)

    input_rows = 0
    blank_rows = 0
    unavailable_input_rows = 0
    total_source_quantity = Decimal("0")
    total_source_cost = Decimal("0")
    try:
        connection.execute("BEGIN")
        for source_row, values in _iter_source_rows(source, plan):
            if not any(_cell(v) for v in values):
                blank_rows += 1
                continue
            input_rows += 1
            quantity = _decimal(values[quantity_idx] if quantity_idx < len(values) else "", "Quantity", source_row)
            cost = _decimal(values[cost_idx] if cost_idx < len(values) else "", "Cost", source_row)
            total_source_quantity += quantity
            total_source_cost += cost

            retained = {
                headers[idx]: _cell(values[idx] if idx < len(values) else "")
                for idx in retained_indices
            }
            identity_available = any(
                _cell(values[idx] if idx < len(values) else "") for idx in identity_indices
            )
            identity_status = (
                "resource_identity_available"
                if identity_available
                else "resource_identity_unavailable"
            )
            if not identity_available:
                unavailable_input_rows += 1

            # The key includes every retained non-date, non-summed field, including
            # prices, currency, subscription/account/region, all resource columns,
            # and every source-specific boundary column. When explicit resource
            # identity is unavailable, these exact fields are the strongest source
            # boundary; the status remains visible and no broader inference occurs.
            key_values = []
            for idx in retained_indices:
                if idx not in {date_idx, quantity_idx, cost_idx}:
                    key_values.append([headers[idx], _cell(values[idx] if idx < len(values) else "")])
            key_json = json.dumps(key_values, ensure_ascii=False, separators=(",", ":"))
            key_hash = hashlib.sha256(key_json.encode("utf-8")).hexdigest()
            current_date_key, current_date = _date_key(values[date_idx] if date_idx < len(values) else "")
            row_json = json.dumps(retained, ensure_ascii=False, separators=(",", ":"))

            connection.execute(
                """
                INSERT INTO aggregates (
                    key_hash, key_json, row_json, quantity, cost,
                    first_date_key, first_date, last_date_key, last_date,
                    source_row_count, identity_status, first_source_row
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                ON CONFLICT(key_hash) DO UPDATE SET
                    quantity = decimal_add(aggregates.quantity, excluded.quantity),
                    cost = decimal_add(aggregates.cost, excluded.cost),
                    first_date = CASE
                        WHEN excluded.first_date_key < aggregates.first_date_key THEN excluded.first_date
                        ELSE aggregates.first_date END,
                    first_date_key = min(aggregates.first_date_key, excluded.first_date_key),
                    last_date = CASE
                        WHEN excluded.last_date_key > aggregates.last_date_key THEN excluded.last_date
                        ELSE aggregates.last_date END,
                    last_date_key = max(aggregates.last_date_key, excluded.last_date_key),
                    source_row_count = aggregates.source_row_count + 1
                """,
                (
                    key_hash, key_json, row_json, _decimal_text(quantity), _decimal_text(cost),
                    current_date_key, current_date, current_date_key, current_date,
                    identity_status, source_row,
                ),
            )
            if progress and input_rows % 10000 == 0:
                progress(input_rows)
            if input_rows % 5000 == 0:
                connection.commit()
                connection.execute("BEGIN")
        connection.commit()

        aggregated_groups = connection.execute(
            "SELECT COUNT(*) FROM aggregates WHERE source_row_count > 1"
        ).fetchone()[0]
        if import_kind == "aws":
            output_audit_columns = (
                AWS_AGGREGATION_AUDIT_COLUMNS
                if aggregated_groups
                else AWS_COMPACT_AUDIT_COLUMNS
            )
        else:
            output_audit_columns = AUDIT_COLUMNS

        output_rows = 0
        unavailable_output_rows = 0
        total_output_quantity = Decimal("0")
        total_output_cost = Decimal("0")
        with output.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(
                fh,
                fieldnames=[*retained_headers, *output_audit_columns],
            )
            writer.writeheader()
            cursor = connection.execute(
                """SELECT row_json, quantity, cost, first_date, last_date,
                          source_row_count, identity_status
                   FROM aggregates ORDER BY first_source_row"""
            )
            quantity_header = headers[quantity_idx]
            cost_header = headers[cost_idx]
            date_header = headers[date_idx]
            for row_json, quantity, cost, first_date, last_date, count, identity_status in cursor:
                row = json.loads(row_json)
                row[quantity_header] = quantity
                row[cost_header] = cost
                row[date_header] = first_date
                audit_values = {
                    "first_usage_date": first_date,
                    "last_usage_date": last_date,
                    "source_row_count": count,
                    "source_file_name": source.name,
                    "source_import_mode": source_import_mode,
                    "resource_identity_status": identity_status,
                }
                row.update({
                    column: audit_values[column]
                    for column in output_audit_columns
                })
                writer.writerow(row)
                output_rows += 1
                total_output_quantity += Decimal(quantity)
                total_output_cost += Decimal(cost)
                if identity_status == "resource_identity_unavailable":
                    unavailable_output_rows += 1

        if total_source_quantity != total_output_quantity or total_source_cost != total_output_cost:
            output.unlink(missing_ok=True)
            raise LargeImportReductionError(
                "Reducer total verification failed; the reduced artifact was discarded. "
                f"Quantity delta={_decimal_text(total_output_quantity - total_source_quantity)}, "
                f"cost delta={_decimal_text(total_output_cost - total_source_cost)}."
            )

        audit = {
            "importKind": import_kind,
            "sourceFileName": source.name,
            "sourceFilePath": str(source),
            "reducedFilePath": str(output),
            "sourceImportMode": source_import_mode,
            "worksheet": plan["sheetName"],
            "headerRow": plan["headerIndex"] + 1,
            "inputRows": input_rows,
            "outputRows": output_rows,
            "blankRowsSkipped": blank_rows,
            "columnsRetained": [*retained_headers, *output_audit_columns],
            "totalSourceQuantity": _decimal_text(total_source_quantity),
            "totalOutputQuantity": _decimal_text(total_output_quantity),
            "totalSourceCost": _decimal_text(total_source_cost),
            "totalOutputCost": _decimal_text(total_output_cost),
            "inputRowsWithUnavailableResourceIdentity": unavailable_input_rows,
            "rowsWithUnavailableResourceIdentity": unavailable_output_rows,
            "aggregationApplied": True,
            "aggregationOccurred": bool(aggregated_groups),
            "aggregatedGroups": aggregated_groups,
            "sourceRowsCollapsed": input_rows - output_rows,
        }
        return _write_audit(output, audit)
    finally:
        connection.close()
        database_path.unlink(missing_ok=True)
        decimal_context.prec = original_decimal_precision


def _reduce_large_onprem(path, output_path, sheet_name="", progress=None):
    """Project a large inventory stream without ever aggregating machine rows."""
    source = Path(path)
    output = Path(output_path)
    plan = preflight_large_import(source, "on_prem", sheet_name)
    headers = plan["headers"]
    retained_indices = plan["retainedIndices"]
    retained_headers = [headers[idx] for idx in retained_indices]
    identity_indices = plan["identityIndices"]
    output.parent.mkdir(parents=True, exist_ok=True)

    input_rows = 0
    blank_rows = 0
    unavailable_rows = 0
    try:
        with output.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=[*retained_headers, *AUDIT_COLUMNS])
            writer.writeheader()
            for source_row, values in _iter_source_rows(source, plan):
                if not any(_cell(value) for value in values):
                    blank_rows += 1
                    continue
                input_rows += 1
                identity_available = any(
                    _cell(values[idx] if idx < len(values) else "")
                    for idx in identity_indices
                )
                identity_status = (
                    "resource_identity_available"
                    if identity_available
                    else "resource_identity_unavailable"
                )
                if not identity_available:
                    unavailable_rows += 1
                row = {
                    headers[idx]: _cell(values[idx] if idx < len(values) else "")
                    for idx in retained_indices
                }
                row.update({
                    "first_usage_date": "",
                    "last_usage_date": "",
                    "source_row_count": 1,
                    "source_file_name": source.name,
                    "source_import_mode": SOURCE_IMPORT_MODES["on_prem"],
                    "resource_identity_status": identity_status,
                })
                writer.writerow(row)
                if progress and input_rows % 10000 == 0:
                    progress(input_rows)
    except Exception:
        output.unlink(missing_ok=True)
        raise

    audit = {
        "importKind": "on_prem",
        "sourceFileName": source.name,
        "sourceFilePath": str(source),
        "reducedFilePath": str(output),
        "sourceImportMode": SOURCE_IMPORT_MODES["on_prem"],
        "worksheet": plan["sheetName"],
        "headerRow": plan["headerIndex"] + 1,
        "inputRows": input_rows,
        "outputRows": input_rows,
        "blankRowsSkipped": blank_rows,
        "columnsRetained": [*retained_headers, *AUDIT_COLUMNS],
        "inputRowsWithUnavailableResourceIdentity": unavailable_rows,
        "rowsWithUnavailableResourceIdentity": unavailable_rows,
        "aggregationApplied": False,
    }
    return _write_audit(output, audit)


def reduce_large_import(path, output_path, import_kind, sheet_name="", progress=None):
    """Preprocess one supported large-file type for the existing parser stack."""
    import_kind = str(import_kind or "").strip().lower()
    if import_kind == "on_prem":
        return _reduce_large_onprem(path, output_path, sheet_name, progress)
    if import_kind in {"aws", "azure"}:
        return _reduce_large_bill(path, output_path, import_kind, sheet_name, progress)
    raise LargeImportReductionError("Large import type must be on_prem, aws, or azure.")
