#!/usr/bin/env python3
"""Regression checks for the Quick BOM's Full BOM-derived visual system."""

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import bom_export  # noqa: E402


def main():
    servers = [
        {
            "name": "Production Web 01",
            "ocpus": 8,
            "memory": 64,
            "disk": 1024,
            "hours": 730,
            "os": "linux",
            "sizeStatus": "ok",
        },
        {
            "name": "Production SQL 01",
            "ocpus": 16,
            "memory": 128,
            "disk": 2048,
            "hours": 730,
            "os": "windows",
            "sizeStatus": "ok",
        },
    ]
    content = bom_export.build_workbook_bytes(
        servers,
        ramp={"months": 60},
        existing_infra_cost=7_000_000,
        bom_name="Quick BOM Style QA",
    )
    wb = load_workbook(BytesIO(content), data_only=False)

    assert wb.sheetnames == ["Overview", "Shape"]
    assert wb.active.title == "Overview"

    bom = wb["Shape"]
    overview = wb["Overview"]
    assert len(bom._images) == 1
    assert len(overview._images) == 1
    assert bom["A2"].value == "OCI Quick BOM | E6 Acceleron"
    assert overview["A2"].value == "Cloud Migration Cost Overview"
    rendered_banner_width = bom._images[0].anchor.ext.width / 9525
    assert round(rendered_banner_width) == bom_export._sheet_pixel_width(bom, 10)
    assert rendered_banner_width > 1300
    assert bom["A2"].fill.fgColor.rgb == "FF2F3437"
    assert bom["J3"].fill.fgColor.rgb == "FF4F5A64"
    assert bom.row_dimensions[3].height == 22
    assert bom["I14"].value == "Cost Summary"
    assert str(bom.merged_cells.ranges).find("I14:J14") >= 0
    assert bom["I15"].value == "Monthly Cost"
    assert bom["I16"].value == "Annual Cost"
    assert bom["I14"].fill.fgColor.rgb == "FF2F3437"
    assert bom["I15"].fill.fgColor.rgb == "FF4F5A64"
    assert bom["J15"].fill.fgColor.rgb == "FFF4F6F8"
    assert bom["J16"].fill.fgColor.rgb == "FFF4F6F8"
    bom_values = [
        cell.value
        for row in bom.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert " Block Volume (Local Storage)" in bom_values
    assert " Boot Volume (Local Storage Sizes)" not in bom_values
    assert overview["A8"].fill.fgColor.rgb == "FF2F3437"
    assert overview["A9"].fill.fgColor.rgb == "FF4F5A64"
    assert overview["B5"].fill.fgColor.rgb == "FFF4F6F8"
    assert bom.sheet_view.showGridLines is False
    assert overview.sheet_view.showGridLines is False

    value_fonts = {
        cell.font.name
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert value_fonts == {"Oracle Sans"}

    formulas = {
        (ws.title, cell.coordinate): cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.data_type == "f"
    }
    assert len(formulas) == 66
    assert formulas[("Shape", "J15")] == "=SUM(G:G)"
    assert formulas[("Shape", "J16")] == "=J15*12"
    assert formulas[("Overview", "B5")] == (
        "='Shape'!J15*(1-$G$5)"
    )
    assert formulas[("Overview", "E5")] == (
        "='Shape'!J16*(1-$G$5)"
    )

    chart = overview._charts[0]
    assert chart.series[0].graphicalProperties.solidFill.srgbClr == "4F5A64"
    assert chart.series[1].graphicalProperties.solidFill.srgbClr == "A7ADB2"
    assert (
        chart._charts[1].series[0]
        .graphicalProperties.line.solidFill.srgbClr
        == "2F3437"
    )

    with_services = load_workbook(
        BytesIO(
            bom_export.build_workbook_bytes(
                servers,
                extra_services=[
                    {
                        "catalogId": "fastconnect",
                        "values": {"speed": "10G", "ports": 1, "__hours": 730},
                    }
                ],
            )
        )
    )
    assert with_services.sheetnames == ["Overview", "Shape"]
    service_values = [
        cell.value
        for row in with_services["Shape"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "Additional OCI Services" in service_values
    assert "B88326" in service_values
    assert not any("Additional Services" == name for name in with_services.sheetnames)

    print("Quick BOM styling regression passed.")


if __name__ == "__main__":
    main()
