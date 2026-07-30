"""Full BOM export - the 12-sheet Oracle BOM deliverable.

Rebuilds the customer-facing BOM workbook (Table of Contents, Assumptions, Rate
Card, Pricing Overview, Compute, Storage, Networking, DR, Security KMS,
Consumption Ramp, Annexure, Applications Migrated to OCI) from the full-fidelity
build-spec (data/bom_template_spec.json), then populates its data sheets from the
app's priced inventory.

The spec carries the reference workbook's visual system: styles, merges, logo,
data-validation, and conditional formatting. Every customer-specific value, assumption,
scenario, server name, formula input, and architecture image is rebuilt from the current
app state or removed.

Compute sheet contract (from the template):
    A VM/Server  B Tier  C Environment  D Master Application  E Master Description
    F Virtual/Physical  G OS Name  H OS Family  I vCPU/Cores  J OCPU (formula)
    K Memory GB  L Storage GB  M Block VPUs (formula)  N Monthly Hours
    O RAM $/mo  P OCPU $/mo  Q Block $/mo  R Total $/mo  S Total $/year
    T Source Shape  U OCI Shape
    Rate Card refs: C8=OCPU rate, C9=RAM rate, C10=block rate, C11=VPU rate,
    C12=default VPUs.  Compute!B9 = E6 optimization factor.
"""

import base64
import datetime
import io
import json
import re
from functools import lru_cache
import shutil
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker, OneCellAnchor,
                                                  TwoCellAnchor)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.formatting.rule import ColorScale, DataBar, FormatObject, Rule
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

SPEC_PATH = Path(__file__).resolve().parent / "data" / "bom_template_spec.json"

# openpyxl uses Pillow ONLY to read an image's width/height when embedding. That single
# dependency kept dropping the architecture diagram (and logos) on machines without Pillow.
# We don't need Pillow for that: PNG dimensions live in the file header. When Pillow is
# missing we parse the header ourselves and monkeypatch openpyxl so images embed anyway -
# the image bytes are written from the original file, never from PIL.
try:
    import PIL  # noqa: F401
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


def _png_dimensions(src):
    """(width, height) from a PNG's IHDR header - no Pillow needed. src is a path or bytes."""
    import struct
    if hasattr(src, "read"):
        pos = src.tell(); head = src.read(24); src.seek(pos)
    elif isinstance(src, (bytes, bytearray)):
        head = bytes(src[:24])
    else:
        with open(src, "rb") as f:
            head = f.read(24)
    if head[:8] != b"\x89PNG\r\n\x1a\n" or head[12:16] != b"IHDR":
        return (0, 0)
    return struct.unpack(">II", head[16:24])


class _HeaderImage:
    """Stand-in for a PIL image. openpyxl's Image class only touches .size, .format, a
    readable .fp, and .close() - never any real decoding - so this is all it needs to embed
    a PNG. The raw bytes are carried through unchanged."""
    def __init__(self, data):
        import io
        self._bytes = bytes(data)
        self.size = _png_dimensions(self._bytes)
        self.format = "PNG"
        self.fp = io.BytesIO(self._bytes)

    def close(self):
        pass


def _enable_pillow_free_images():
    """Let openpyxl embed PNGs without Pillow by supplying dimensions from the header and
    handing back the raw bytes. openpyxl re-fetches the image via _import_image at write
    time, so this covers both metadata and the actual save. No-op when Pillow is present."""
    if HAS_PILLOW:
        return
    try:
        from openpyxl.drawing import image as _oxml_image
    except Exception:
        return

    def _import_image(img):
        if isinstance(img, _HeaderImage):
            return img
        if hasattr(img, "read"):
            pos = img.tell(); data = img.read(); img.seek(pos)
        else:
            with open(img, "rb") as f:
                data = f.read()
        return _HeaderImage(data)

    _oxml_image._import_image = _import_image
    # The constructor raises ImportError unless PILImage is truthy.
    if not getattr(_oxml_image, "PILImage", None):
        _oxml_image.PILImage = _HeaderImage


_enable_pillow_free_images()

COMPUTE_SHEET = "Compute"
COMPUTE_HEADER_ROW = 13
COMPUTE_FIRST_ROW = 14
COMPUTE_LAST_TEMPLATE_ROW = 689          # rows the reference deliverable ships with
COMPUTE_FORMULA_COLS = ["J", "M", "O", "P", "Q", "R", "S"]
COMPUTE_ALL_COLS = list("ABCDEFGHIJKLMNOPQRSTU")

APPS_SHEET = "Applications Migrated to OCI"
APPS_FIRST_ROW = 7
APPS_LAST_TEMPLATE_ROW = 207
APPS_FORMULA_COLS = list("BCDEFGH")

STORAGE_SHEET = "Storage"
STORAGE_HEADER_ROW = 10
STORAGE_FIRST_ROW = 11
STORAGE_LAST_TEMPLATE_ROW = 21


# ---------------------------------------------------------------------------
# Spec -> workbook (full-fidelity rebuild)
# ---------------------------------------------------------------------------
def _color(d):
    if not d:
        return None
    if "rgb" in d:
        return Color(rgb=d["rgb"])
    if "theme" in d:
        return Color(theme=d["theme"], tint=d.get("tint", 0.0))
    if "indexed" in d:
        return Color(indexed=d["indexed"])
    return None


def _font(d):
    if not d:
        return None
    return Font(name=d.get("name"), sz=d.get("size"), bold=d.get("bold", False),
                italic=d.get("italic", False), underline=d.get("underline"),
                strike=d.get("strike", False), color=_color(d.get("color")))


def _fill(d):
    if not d:
        return None
    return PatternFill(patternType=d["pattern"],
                       fgColor=_color(d.get("fgColor")) or Color(),
                       bgColor=_color(d.get("bgColor")) or Color())


def _side(d):
    return Side(style=d["style"], color=_color(d.get("color"))) if d else None


def _border(d):
    if not d:
        return None
    return Border(left=_side(d.get("left")), right=_side(d.get("right")),
                  top=_side(d.get("top")), bottom=_side(d.get("bottom")),
                  diagonal=_side(d.get("diagonal")),
                  diagonalUp=d.get("diagonalUp", False),
                  diagonalDown=d.get("diagonalDown", False))


def _align(d):
    if not d:
        return None
    return Alignment(horizontal=d.get("horizontal"), vertical=d.get("vertical"),
                     textRotation=d.get("textRotation", 0),
                     wrapText=d.get("wrapText", False),
                     shrinkToFit=d.get("shrinkToFit", False), indent=d.get("indent", 0))


def _unjval(v):
    if isinstance(v, dict) and "_type" in v:
        if v["_type"] == "datetime":
            return datetime.datetime.fromisoformat(v["iso"])
        if v["_type"] == "date":
            return datetime.date.fromisoformat(v["iso"])
        if v["_type"] == "time":
            return datetime.time.fromisoformat(v["iso"])
    return v


def load_spec():
    return json.loads(SPEC_PATH.read_text())


def build_workbook(spec):
    """Rebuild the template workbook from the build-spec (styles, images, formulas)."""
    styles = {}
    for sid, d in spec["styles"].items():
        styles[sid] = {
            "font": _font(d.get("font")), "fill": _fill(d.get("fill")),
            "border": _border(d.get("border")), "alignment": _align(d.get("alignment")),
            "number_format": d.get("number_format"),
        }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sd in spec["sheets"]:
        ws = wb.create_sheet(sd["name"])
        ws.sheet_state = sd.get("state", "visible")
        if sd.get("tab_color"):
            ws.sheet_properties.tabColor = _color(sd["tab_color"])
        v = sd.get("view", {})
        ws.sheet_view.showGridLines = v.get("showGridLines", True)
        ws.sheet_view.zoomScale = v.get("zoomScale", 100)
        if v.get("showRowColHeaders") is False:
            ws.sheet_view.showRowColHeaders = False
        if sd.get("freeze_panes"):
            ws.freeze_panes = sd["freeze_panes"]
        sf = sd.get("sheet_format", {})
        if sf.get("defaultColWidth"):
            ws.sheet_format.defaultColWidth = sf["defaultColWidth"]
        if sf.get("defaultRowHeight"):
            ws.sheet_format.defaultRowHeight = sf["defaultRowHeight"]

        for col, d in sd.get("column_dimensions", {}).items():
            cd = ws.column_dimensions[col]
            if "width" in d:
                cd.width = d["width"]
            if d.get("hidden"):
                cd.hidden = True
            if d.get("outlineLevel"):
                cd.outlineLevel = d["outlineLevel"]
            if "range" in d:
                cd.min, cd.max = d["range"]
        for row, d in sd.get("row_dimensions", {}).items():
            rd = ws.row_dimensions[int(row)]
            if "height" in d:
                rd.height = d["height"]
            if d.get("hidden"):
                rd.hidden = True
            if d.get("outlineLevel"):
                rd.outlineLevel = d["outlineLevel"]

        # Merges first: merge_cells() wipes styles set inside the range.
        for rng in sd.get("merged_cells", []):
            ws.merge_cells(rng)

        for coord, cd in sd["cells"].items():
            c = ws[coord]
            if "f" in cd:
                c.value = cd["f"]
            elif "v" in cd:
                c.value = _unjval(cd["v"])
            if "s" in cd:
                st = styles[cd["s"]]
                if st["font"]:
                    c.font = st["font"]
                if st["fill"]:
                    c.fill = st["fill"]
                if st["border"]:
                    c.border = st["border"]
                if st["alignment"]:
                    c.alignment = st["alignment"]
                if st["number_format"]:
                    c.number_format = st["number_format"]
            if "hyperlink" in cd:
                from openpyxl.worksheet.hyperlink import Hyperlink
                hl = cd["hyperlink"]
                c.hyperlink = Hyperlink(ref=coord, target=hl.get("target"),
                                        location=hl.get("location"), tooltip=hl.get("tooltip"))
            if "comment" in cd:
                c.comment = Comment(cd["comment"]["text"], cd["comment"].get("author") or "")

        if sd.get("auto_filter"):
            ws.auto_filter.ref = sd["auto_filter"]

        for r in sd.get("conditional_formatting", []):
            kw = {"type": r["type"]}
            if "operator" in r:
                kw["operator"] = r["operator"]
            if "formula" in r:
                kw["formula"] = r["formula"]
            if "priority" in r:
                kw["priority"] = r["priority"]
            if r.get("stopIfTrue"):
                kw["stopIfTrue"] = True
            if "dxf" in r:
                dx = r["dxf"]
                kw["dxf"] = DifferentialStyle(font=_font(dx.get("font")),
                                              fill=_fill(dx.get("fill")),
                                              border=_border(dx.get("border")))
            if "colorScale" in r:
                cs = r["colorScale"]
                kw["colorScale"] = ColorScale(
                    cfvo=[FormatObject(type=o["type"], val=o["val"]) for o in cs["cfvo"]],
                    color=[_color(c) for c in cs["colors"]])
            if "dataBar" in r:
                db = r["dataBar"]
                kw["dataBar"] = DataBar(
                    cfvo=[FormatObject(type=db.get("minType", "min"), val=db.get("minVal")),
                          FormatObject(type=db.get("maxType", "max"), val=db.get("maxVal"))],
                    color=_color(db.get("color")),
                    showValue=db.get("showValue", True),
                    minLength=db.get("minLength"),
                    maxLength=db.get("maxLength"))
            # A rule whose type needs a payload we didn't build would serialise as a bare
            # <cfRule type="dataBar"/>, which Excel rejects outright - it opens the file with
            # "we found a problem with some content" and strips the sheet. openpyxl and
            # LibreOffice both ignore the empty rule, so it only ever showed up in Excel.
            # Better to drop a rule we can't express than to ship an unopenable workbook.
            _payload = {"dataBar": "dataBar", "colorScale": "colorScale", "iconSet": "iconSet"}
            _needs = _payload.get(r["type"])
            if _needs and _needs not in kw:
                continue
            ws.conditional_formatting.add(r["range"], Rule(**kw))

        for d in sd.get("data_validations", []):
            dv = DataValidation(type=d.get("type"), operator=d.get("operator"),
                                formula1=d.get("formula1"), formula2=d.get("formula2"),
                                allowBlank=d.get("allowBlank", False),
                                showInputMessage=d.get("showInputMessage", False),
                                showErrorMessage=d.get("showErrorMessage", False),
                                errorTitle=d.get("errorTitle"), error=d.get("error"),
                                promptTitle=d.get("promptTitle"), prompt=d.get("prompt"))
            for rng in d["ranges"].split():
                dv.add(rng)
            ws.add_data_validation(dv)

        for im in sd.get("images", []):
            data = base64.b64decode(spec["images"][im["image_ref"]]["base64"])
            img = XLImage(io.BytesIO(data))
            a = im["anchor"]
            fr = a["from"]
            m1 = AnchorMarker(col=fr["col"], colOff=fr["colOff"], row=fr["row"], rowOff=fr["rowOff"])
            if a["type"] == "TwoCellAnchor" and "to" in a:
                t = a["to"]
                m2 = AnchorMarker(col=t["col"], colOff=t["colOff"], row=t["row"], rowOff=t["rowOff"])
                img.anchor = TwoCellAnchor(_from=m1, to=m2)
            else:
                ext = a.get("ext_emu")
                img.anchor = OneCellAnchor(
                    _from=m1, ext=XDRPositiveSize2D(cx=ext["cx"], cy=ext["cy"]) if ext else None)
            ws.add_image(img)

    order = spec["workbook"]["sheet_order"]
    # Open on the front of the deliverable, not wherever the source workbook was last saved
    # (the template had "Compute" active, so the export opened mid-way through the BOM).
    landing = "Table of Contents" if "Table of Contents" in order else order[0]
    wb.active = order.index(landing)
    for i, ws in enumerate(wb.worksheets):
        ws.sheet_view.tabSelected = (i == wb.active)
    return wb


def _postprocess(path, spec):
    """Restore the theme, the default font, and zero-height rows (openpyxl drops them)."""
    wbk = spec["workbook"]

    def font_xml(d):
        parts = [f'<name val="{d["name"]}"/>']
        if "size" in d:
            parts.append(f'<sz val="{d["size"]:g}"/>')
        if d.get("bold"):
            parts.append("<b/>")
        if d.get("italic"):
            parts.append("<i/>")
        c = d.get("color")
        if c and "rgb" in c:
            parts.append(f'<color rgb="{c["rgb"]}"/>')
        elif c and "theme" in c:
            parts.append(f'<color theme="{c["theme"]}"/>')
        return "<font>" + "".join(parts) + "</font>"

    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(path) as zin:
        # New service tabs are inserted among the reference sheets, so worksheet XML
        # numbers no longer match the original 12-sheet spec. Resolve each sheet by its
        # workbook relationship instead of assuming sheet1/sheet2/... positional identity.
        import xml.etree.ElementTree as ET
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        book_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        targets = {
            rel.get("Id"): rel.get("Target")
            for rel in rel_root.findall(f"{{{rel_ns}}}Relationship")
        }
        sheet_paths = {}
        for sheet in book_root.findall(f".//{{{main_ns}}}sheet"):
            target = targets.get(sheet.get(rel_attr), "")
            if target.startswith("/"):
                target = target.lstrip("/")
            elif target and not target.startswith("xl/"):
                target = f"xl/{target}"
            sheet_paths[sheet.get("name")] = target

        zero_rows = {}
        for sd in spec["sheets"]:
            zr = [r for r, d in sd.get("row_dimensions", {}).items()
                  if d.get("height") == 0]
            target = sheet_paths.get(sd["name"])
            if zr and target:
                zero_rows[target] = zr

    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/theme/theme1.xml" and wbk.get("theme_xml_base64"):
                data = base64.b64decode(wbk["theme_xml_base64"])
            elif item.filename == "xl/styles.xml" and wbk.get("default_font"):
                txt = data.decode("utf-8")
                txt = re.sub(r"(<fonts[^>]*>)<font>.*?</font>",
                             lambda m: m.group(1) + font_xml(wbk["default_font"]), txt, count=1)
                data = txt.encode("utf-8")
            elif item.filename in zero_rows:
                txt = data.decode("utf-8")
                for r in zero_rows[item.filename]:
                    m = re.search(f'<row r="{r}"[^>]*>', txt)
                    if not m:
                        continue
                    tag = m.group(0)
                    if ' ht="' not in tag:
                        new = tag.replace(f'<row r="{r}"', f'<row r="{r}" ht="0"', 1)
                        if "customHeight=" not in new:
                            new = new.replace(f'<row r="{r}"', f'<row r="{r}" customHeight="1"', 1)
                        txt = txt.replace(tag, new, 1)
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)


# ---------------------------------------------------------------------------
# Field resolution from the app's uploaded inventory
# ---------------------------------------------------------------------------
def _norm(v):
    return re.sub(r"[^a-z0-9]+", " ", str(v or "").lower()).strip()


def _clean(v):
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return "" if s.lower() in {"nan", "none", "nat"} else s


def _field_source_text(f):
    """The field's ORIGINAL column header. The app renames CPU/memory columns to
    'Application Details: OCPUs' / '... Memory per server (GB)', so matching on the
    renamed label would wrongly resolve e.g. 'Application' to the OCPU column."""
    original = f.get("cpuSourceLabel") or f.get("memorySourceLabel") or f.get("sourceHeader")
    return _norm(original or f.get("label"))


def _find_field(fields, *needle_sets):
    """First field whose ORIGINAL header matches any needle set (all terms present as
    WHOLE WORDS). Word-boundary matching, not substring, so e.g. the 'app' role can't
    latch onto the 'Mapping confidence' column ('app' is a substring of 'mapping') and
    name every segment after the mapping-confidence value ('Service guide')."""
    for needles in needle_sets:
        for f in fields or []:
            if not isinstance(f, dict):
                continue
            tokens = set(_field_source_text(f).split())
            if all(n in tokens for n in needles):
                return f.get("key")
    return None


def _distinct_sites(fields, rows):
    """How many distinct sites the inventory actually names - None when it has no site
    column at all (most VM exports don't). Drives whether the diagram draws remote sites."""
    key = _find_field(fields, ["site"], ["location"], ["datacenter"], ["data", "center"],
                      ["campus"], ["facility"])
    if not key:
        return None
    vals = {_clean(r.get(key)) for r in (rows or []) if _clean(r.get(key))}
    return len(vals) or None


def _resolve_inventory_keys(fields):
    return {
        "server": _find_field(fields, ["server", "name"], ["vm", "name"], ["host", "name"], ["machine"]),
        "os_name": _find_field(fields, ["guest", "os"], ["os", "name"], ["operating", "system"]),
        "os_family": _find_field(fields, ["os", "type"], ["os", "family"], ["platform"]),
        "virt": _find_field(fields, ["physical"], ["virtual"], ["server", "type"]),
        "env": _find_field(fields, ["environment"], ["env"]),
        "app": _find_field(fields, ["application", "name"], ["application"], ["app"]),
        "tier": _find_field(fields, ["tier"]),
        "desc": _find_field(fields, ["description"]),
    }


# ---------------------------------------------------------------------------
# Populate
# ---------------------------------------------------------------------------
def _clear_range(ws, first_row, last_row, cols):
    for r in range(first_row, last_row + 1):
        for c in cols:
            ws[f"{c}{r}"] = None


def _populate_compute(ws, servers, hours, rate_refs=None, shape_label=""):
    """Write the app's servers into the Compute sheet, translating the template's
    row-14 formulas down to each row and clearing any unused template rows. The live
    OCPU/RAM/block/VPU formulas point at the Rate Card cells that the rebuilt (used-only)
    rate card actually placed those rates on, so the math is transparent."""
    rate_refs = rate_refs or {}
    protos = {c: ws[f"{c}{COMPUTE_FIRST_ROW}"].value for c in COMPUTE_FORMULA_COLS}
    proto_styles = {c: ws[f"{c}{COMPUTE_FIRST_ROW}"]._style for c in COMPUTE_ALL_COLS}
    # The app prices each line item to the cent and then sums. Mirror that rounding here
    # (same rates/refs, just ROUND-ed) so the workbook ties out to the app to the penny
    # instead of drifting by fractions of a cent across hundreds of rows.
    R = COMPUTE_FIRST_ROW
    ram_ref = rate_refs.get("ram")
    ocpu_ref = rate_refs.get("ocpu")
    block_ref = rate_refs.get("block")
    vpu_ref = rate_refs.get("vpu")
    vpus_ref = rate_refs.get("vpus")
    # IMPORTANT: only OVERRIDE a proto when we have the matching Rate Card row. Setting it to 0
    # made `if proto:` below skip the write, which left every row past the template's last
    # shipped row (COMPUTE_LAST_TEMPLATE_ROW) with NO cost formula at all - silently understating
    # any inventory with more servers than the template ships rows for. Keeping the template's own
    # row-14 formula as the fallback guarantees every data row gets priced.
    protos["O"] = (f"=ROUND(K{R}*N{R}*'Rate Card'!$C${ram_ref},2)" if ram_ref else protos.get("O"))
    # Compute optimization adjusts the OCPU/RAM QUANTITIES (columns J and K) - it must NOT
    # also discount the price here, or it would double-count. Price flows straight from the
    # optimized sizing.
    protos["P"] = (f"=ROUND(J{R}*N{R}*'Rate Card'!$C${ocpu_ref},2)" if ocpu_ref else protos.get("P"))
    if block_ref and vpu_ref:
        protos["Q"] = (f"=IF($L{R}=\"\",0,ROUND($L{R}*'Rate Card'!$C${block_ref},2)"
                       f"+ROUND($L{R}*'Rate Card'!$C${vpu_ref}*IF($M{R}=\"\",0,$M{R}),2))")
    # Block-VPU seed (column M) references the Rate Card default-VPUs cell.
    protos["M"] = (f"=IF($A{R}<>\"\",'Rate Card'!$C${vpus_ref},\"\")" if vpus_ref else '=""')

    # The Polaris template reserves S for Total Annual. T/U show what each server maps
    # FROM (source) and TO (OCI) without overwriting that annual-cost formula.
    from copy import copy as _copy_style
    _hdr_style = ws.cell(13, 1)._style
    ws.cell(13, 20).value = "Source Shape (Mapped From)"
    ws.cell(13, 20)._style = _copy_style(_hdr_style)
    ws.cell(13, 21).value = "OCI Shape (Mapped To)"
    ws.cell(13, 21)._style = _copy_style(_hdr_style)
    ws.column_dimensions["T"].width = 24
    ws.column_dimensions["U"].width = 22
    # The reference headers hard-coded E6. Always restate them from this export.
    target_label = shape_label or "OCI Flex"
    ws["O13"] = f"{target_label} RAM Monthly"
    ws["P13"] = f"{target_label} OCPU Monthly"
    ws["R13"] = f"{target_label} Total Monthly"
    ws["S13"] = f"{target_label} Total Annual"
    ws["F5"] = "Compute rows and pricing below are generated from the current estimate."
    ws["F6"] = (
        "VM-attached block storage uses the capacity and performance-unit rates listed "
        "in this export's Rate Card."
    )
    ws["F7"] = (
        "OCPU quantities use the source CPU interpretation selected in the estimator."
    )
    ws["F8"] = (
        "Monthly hours are carried from the current estimate for each workload."
    )
    ws["F9"] = (
        "Rightsizing is reflected only when it is enabled in the current estimate."
    )
    ws["F10"] = (
        "Software licensing is included only when selected or detected by the estimator."
    )

    last_written = COMPUTE_FIRST_ROW - 1
    for i, s in enumerate(servers):
        r = COMPUTE_FIRST_ROW + i
        # Extend styling past the template's shipped rows if the inventory is bigger.
        if r > COMPUTE_LAST_TEMPLATE_ROW:
            for c in COMPUTE_ALL_COLS:
                ws[f"{c}{r}"]._style = proto_styles[c]
        ws[f"A{r}"] = s.get("server") or s.get("app") or f"Server {i + 1}"
        ws[f"B{r}"] = s.get("tier") or None
        ws[f"C{r}"] = s.get("env") or None
        ws[f"D{r}"] = s.get("app") or None
        ws[f"E{r}"] = s.get("desc") or None
        ws[f"F{r}"] = s.get("virt") or None
        ws[f"G{r}"] = s.get("os_name") or None
        ws[f"H{r}"] = s.get("os_family") or None
        ws[f"I{r}"] = s.get("vcpu") or None
        ws[f"K{r}"] = s.get("memory_gb") or None
        ws[f"L{r}"] = s.get("storage_gb") or None
        # Per-row hours from the data source (falls back to the global hours).
        ws[f"N{r}"] = s.get("hours") or hours
        ws[f"T{r}"] = s.get("source_shape") or None          # Mapped From (source)
        ws[f"U{r}"] = s.get("shape") or shape_label or None  # Mapped To (OCI)
        ws[f"T{r}"]._style = _copy_style(proto_styles["A"])
        ws[f"U{r}"]._style = _copy_style(proto_styles["A"])
        for c in COMPUTE_FORMULA_COLS:
            proto = protos.get(c)
            if proto:
                ws[f"{c}{r}"] = Translator(
                    proto, origin=f"{c}{COMPUTE_FIRST_ROW}").translate_formula(f"{c}{r}")
        # The source workbook shipped its data rows COLLAPSED (hidden=True on all 1,061 of
        # them), and the spec reproduces that faithfully - so every server we wrote was
        # invisible in the export and the sheet looked empty. Show the rows we populate.
        rd = ws.row_dimensions[r]
        rd.hidden = False
        if not rd.height:
            rd.height = 28.0
        last_written = r

    # Blank out any leftover reference rows so SUBTOTAL/COUNTA don't pick them up, and
    # collapse them so the sheet ends at the last real server instead of trailing a
    # thousand empty rows.
    if last_written < COMPUTE_LAST_TEMPLATE_ROW:
        _clear_range(ws, last_written + 1, COMPUTE_LAST_TEMPLATE_ROW, COMPUTE_ALL_COLS)
        for r in range(last_written + 1, COMPUTE_LAST_TEMPLATE_ROW + 1):
            ws.row_dimensions[r].hidden = True
    return last_written


def _populate_apps(ws, apps, servers=None, shape_label=""):
    """Applications Migrated to OCI: current app names, formulas, and example VMs only."""
    protos = {c: ws[f"{c}{APPS_FIRST_ROW}"].value for c in APPS_FORMULA_COLS}
    proto_styles = {c: ws[f"{c}{APPS_FIRST_ROW}"]._style for c in "ABCDEFGHI"}
    _clear_range(ws, APPS_FIRST_ROW, APPS_LAST_TEMPLATE_ROW, list("ABCDEFGHI"))
    target_label = shape_label or "OCI Flex"
    ws["C6"] = f"{target_label} Monthly Baseline"
    ws["D6"] = f"{target_label} Annual Baseline"
    ws["I6"] = "Example VMs"
    by_app = {}
    for server in servers or []:
        app = _clean(server.get("app"))
        name = _clean(server.get("server"))
        if app and name:
            by_app.setdefault(app, []).append(name)
    last = APPS_FIRST_ROW - 1
    for i, name in enumerate(apps):
        r = APPS_FIRST_ROW + i
        if r > APPS_LAST_TEMPLATE_ROW:
            break
        for c in "ABCDEFGHI":
            ws[f"{c}{r}"]._style = proto_styles[c]
        ws.row_dimensions[r].hidden = False
        ws[f"A{r}"] = name
        for c in APPS_FORMULA_COLS:
            proto = protos.get(c)
            if proto:
                ws[f"{c}{r}"] = Translator(
                    proto, origin=f"{c}{APPS_FIRST_ROW}").translate_formula(f"{c}{r}")
        for c in "BCDEFGH":
            ws[f"{c}{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{r}"].number_format = "#,##0"
        for c in "CD":
            ws[f"{c}{r}"].number_format = '"$"#,##0.00'
        for c in "EFGH":
            ws[f"{c}{r}"].number_format = "#,##0.00"
        ws[f"I{r}"] = ", ".join(by_app.get(name, [])[:6]) or None
        last = r
    if last < APPS_LAST_TEMPLATE_ROW:
        _clear_range(ws, last + 1, APPS_LAST_TEMPLATE_ROW, list("ABCDEFGHI"))
        for r in range(last + 1, APPS_LAST_TEMPLATE_ROW + 1):
            ws.row_dimensions[r].hidden = True


def _populate_storage(ws, storage_rows, rate_refs=None, file_rate=None):
    """Storage sheet: object/file-storage candidates (cleared when the app has none).

    Each row can supply either (gb, rate) -> I = gb*rate, or a direct `monthly` value when
    capacity isn't known (cloud-bill storage is usage-priced, so GB isn't always available).
    File-storage rows point their rate cell (H) at the Rate Card file-storage rate for
    transparency."""
    rate_refs = rate_refs or {}
    file_ref = rate_refs.get("file")
    proto = ws[f"A{STORAGE_FIRST_ROW}"]._style
    headers = [
        "Workload / Service", "Tier", "Environment", "Application", "Source Signal",
        "OCI Storage Target", "Capacity (GB)", "Unit Rate", "Monthly", "Annual",
    ]
    ws["F5"] = "Rows below are populated from storage services in the current estimate."
    ws["A6"] = "Includes only storage entries present in the current estimator state."
    ws["F6"] = "VM-attached block storage remains priced on the Compute tab."
    ws["F7"] = "No capacities or storage scenarios are imported from the styling reference."
    ws["A8"] = "Storage detail below is generated from the current estimate."
    for c, value in enumerate(headers, 1):
        ws.cell(STORAGE_HEADER_ROW, c).value = value
    _clear_range(ws, STORAGE_FIRST_ROW, STORAGE_LAST_TEMPLATE_ROW, list("ABCDEFGHIJ"))
    for i, s in enumerate(storage_rows):
        r = STORAGE_FIRST_ROW + i
        if r > STORAGE_LAST_TEMPLATE_ROW:
            for c in "ABCDEFGHIJ":
                ws[f"{c}{r}"]._style = proto
        ws[f"A{r}"] = s.get("server")
        ws[f"B{r}"] = s.get("tier")
        ws[f"C{r}"] = s.get("env")
        ws[f"D{r}"] = s.get("app")
        ws[f"E{r}"] = s.get("signal")
        ws[f"F{r}"] = s.get("target")
        ws[f"G{r}"] = s.get("gb")
        rate = s.get("rate")
        if s.get("gb") and rate:
            if file_ref and file_rate and abs(float(rate) - float(file_rate)) < 1e-9:
                ws[f"H{r}"] = f"='Rate Card'!$C${file_ref}"
            else:
                ws[f"H{r}"] = rate
            ws[f"I{r}"] = f"=G{r}*H{r}"
        else:
            ws[f"H{r}"] = rate
            ws[f"I{r}"] = round(float(s.get("monthly") or 0), 2)
        ws[f"J{r}"] = f"=I{r}*12"


def _cloud_storage_rows(pricing):
    """Aggregate the cloud-bill Storage-category services by OCI product for the Storage
    sheet, so it itemizes what rolled into the Pricing Overview Storage line."""
    import bom_export
    agg = {}
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        if bom_export._cloud_product_group(r.get("ociServiceCategory"),
                                           r.get("sourceService")) != "Storage":
            continue
        prod = _clean(r.get("ociProduct")) or "OCI Storage"
        specs = r.get("specs") or {}
        gb = (float(specs.get("blockStorageGb") or 0) + float(specs.get("fileStorageGb") or 0)
              + float(specs.get("cloudStorageGb") or 0))
        a = agg.setdefault(prod, {"gb": 0.0, "monthly": 0.0, "svc": _clean(r.get("sourceService"))})
        a["gb"] += gb
        a["monthly"] += float(r.get("monthly") or 0)
    rows = []
    for prod, v in sorted(agg.items(), key=lambda kv: -kv[1]["monthly"]):
        if v["monthly"] <= 0:
            continue
        rows.append({
            "server": prod, "signal": "Mapped from cloud bill",
            "target": prod, "app": v["svc"],
            "gb": round(v["gb"], 2) or None,
            "rate": round(v["monthly"] / v["gb"], 6) if v["gb"] else None,
            "monthly": round(v["monthly"], 2),
        })
    return rows


def _extra_storage_rows(extra_priced):
    """Summarize added Storage-category services in the reference Storage table."""
    rows = []
    for service in extra_priced or []:
        if _service_category(service.get("group"), service.get("name")) != "Storage":
            continue
        qty = float(service.get("qty") or 0)
        unit = _norm(service.get("unit"))
        gb = qty if "gb" in unit else None
        rows.append({
            "server": service.get("name"),
            "signal": "Added in service catalog",
            "target": service.get("name"),
            "gb": round(gb, 2) if gb else None,
            "rate": float(service.get("rate") or 0) or None,
            "monthly": round(float(service.get("monthly") or 0), 2),
        })
    return rows


RATE_CARD_HDR_ROW = 7      # header row for the rate table
RATE_CARD_FIRST_ROW = 8    # first data row
RATE_CARD_CLEAR_LAST = 60  # scrub the template's fixed sections well past the last row


def _collect_rate_card_entries(shape, block_rate, vpu_rate, default_vpus, hours,
                               file_rate, windows_rate, windows_sku, windows_priced,
                               servers, storage_rows, pricing, extra_services,
                               is_cloud_bill):
    """Build the list of rate-card lines that were ACTUALLY used in this build. Core
    compute/storage/licensing inputs (carrying a `key` so the sheets can reference them),
    plus every distinct mapped-service SKU that appears in the priced line items and any
    user-added OCI services. Returns a list of {sku, name, val, unit, note, key}."""
    entries = []

    def add(sku, name, val, unit, note, key=None):
        entries.append({"sku": (sku or "N/A"), "name": name, "val": val,
                        "unit": unit, "note": note, "key": key})

    has_compute = bool(servers)
    has_block = any(s.get("storage_gb") for s in servers)
    has_file = any("file" in (s.get("target") or "").lower() for s in (storage_rows or []))

    label = (shape or {}).get("shortLabel") or (shape or {}).get("label") or "OCI"
    if has_compute and shape:
        if shape.get("computeRate") is not None:
            add(shape.get("computeSku"), f"{label} OCPU per hour", float(shape["computeRate"]),
                "per OCPU-hour", f"OCI {label} Flex Compute OCPU pricing (Compute col P).", "ocpu")
        if shape.get("memoryRate") is not None:
            add(shape.get("memorySku"), f"{label} RAM GB per hour", float(shape["memoryRate"]),
                "per GB-hour", f"OCI {label} Flex Compute memory pricing (Compute col O).", "ram")
    if has_block:
        add("B91961", "VM Block Volume Storage", float(block_rate), "per GB-month",
            "VM-attached block storage (Compute col Q).", "block")
        add("B91962", "VM Block Volume Performance Units", float(vpu_rate), "per GB-month",
            "Block Volume performance (VPU) component (Compute col Q).", "vpu")
        add("N/A", "Default VM Block Volume VPUs", float(default_vpus), "VPUs / GB-month",
            "Default performance units seeded into the Compute VPU column (col M).", "vpus")
    if has_compute:
        add("N/A", "Monthly hours", float(hours), "hours / month",
            "Default monthly-hours assumption (Compute col N).", "hours")
        if not is_cloud_bill:
            add("N/A", "vCPU / core to OCPU conversion",
                "Virtual: 2 vCPU = 1 OCPU; Physical: 1 core = 1 OCPU", "conversion",
                "Used when deriving OCPUs from the source inventory.", "conv")
    if has_file:
        add("B89057", "File Storage Service", float(file_rate), "per GB-month",
            "File / NAS storage rows (Storage sheet).", "file")
    if windows_priced and windows_rate is not None:
        add(windows_sku, "Windows OS licensing", float(windows_rate), "per OCPU-hour",
            "Windows rows; shown as 3rd Party Licensing on the Pricing Overview.", "windows")

    # Every distinct mapped-service SKU that appears in the priced line items (covers
    # cloud-bill services: OIC, ADW, Load Balancer, WAF, DNS, Object Storage, etc.). A SKU
    # can appear on many lines (e.g. OIC's single priced anchor plus zero-cost consolidated
    # rows) - keep the line with the highest unit rate so the rate card shows the real rate.
    core_skus = {e["sku"] for e in entries if e["sku"] and e["sku"] != "N/A"}
    info = {}

    def note_used(sku, name, unit, val):
        if not sku or sku in core_skus:
            return
        val = val if isinstance(val, (int, float)) else None
        cur = info.get(sku)
        if cur is None:
            info[sku] = {"name": name or sku, "unit": unit or "", "val": val}
        elif (val or 0) > (cur["val"] or 0):
            cur.update(name=name or cur["name"], unit=unit or cur["unit"], val=val)

    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        for li in (r.get("lineItems") or []):
            note_used(_clean(li.get("sku")), _clean(li.get("description")),
                      _clean(li.get("unit")), li.get("rate"))

    if extra_services:
        try:
            import oci_catalog
            priced, _ = oci_catalog.price_extras(extra_services, hours)
            for service in priced:
                service_lines = service.get("skus") or [{
                    "sku": service.get("sku"),
                    "desc": service.get("name"),
                    "rate": service.get("rate"),
                }]
                for line in service_lines:
                    note_used(
                        _clean(line.get("sku")),
                        _clean(line.get("desc")) or _clean(service.get("name")),
                        _clean(service.get("unit")),
                        line.get("rate"),
                    )
        except Exception:
            pass

    for sku, v in info.items():
        add(sku, v["name"], v["val"], v["unit"], "Mapped-service rate used in this build.")

    return entries


def _write_rate_card(ws, entries):
    """Rewrite the Rate Card sheet with only the used entries, sorted alphabetically by
    SKU (N/A inputs after the real SKUs, by name). Returns {key: row} so the other sheets
    can point their live formulas at the exact cells for transparency."""
    # Preserve the template's header + data cell styling, then scrub its fixed sections.
    cols = "ABCDE"
    hdr_style = {c: ws.cell(RATE_CARD_HDR_ROW, i)._style for i, c in enumerate(cols, 1)}
    data_style = {c: ws.cell(RATE_CARD_FIRST_ROW, i)._style for i, c in enumerate(cols, 1)}
    # The template's section headers span merged cells; unmerge anything in the scrub
    # region so we can overwrite it, then blank it out.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 6 and mr.max_row <= RATE_CARD_CLEAR_LAST:
            ws.unmerge_cells(str(mr))
    for r in range(6, RATE_CARD_CLEAR_LAST + 1):
        for i in range(1, 7):
            ws.cell(r, i).value = None

    ws.cell(4, 1).value = "RATE CARD - ONLY THE SKUS / RATES USED IN THIS BUILD"
    headers = ["SKU", "Input", "Value", "Unit", "Workbook Use / Note"]
    for i, (c, h) in enumerate(zip(cols, headers), 1):
        cell = ws.cell(RATE_CARD_HDR_ROW, i)
        cell.value = h
        cell._style = hdr_style[c]

    def sort_key(e):
        sku = e["sku"]
        if sku and sku != "N/A":
            return (0, sku.upper(), e["name"].upper())
        return (1, "", e["name"].upper())

    refs = {}
    for idx, e in enumerate(sorted(entries, key=sort_key)):
        r = RATE_CARD_FIRST_ROW + idx
        vals = [e["sku"], e["name"], e["val"], e["unit"], e["note"]]
        for i, (c, v) in enumerate(zip(cols, vals), 1):
            cell = ws.cell(r, i)
            cell.value = v
            cell._style = data_style[c]
        cval = e["val"]
        if isinstance(cval, (int, float)):
            ws.cell(r, 3).number_format = "#,##0" if float(cval).is_integer() else "0.0000"
        if e.get("key"):
            refs[e["key"]] = r
        else:
            # The OCPU / memory rates usually arrive through the mapped-service path (from the
            # priced line items) rather than the keyed compute path, so they carry no key. Bind
            # them by unit so the Compute formulas can point at the row this build actually put
            # them on, instead of relying on them happening to sort into the template's
            # hard-coded $C$8 / $C$9.
            _u = _norm(e.get("unit"))
            if _u in {"ocpu-hour", "per ocpu-hour"} and "ocpu" not in refs:
                refs["ocpu"] = r
            elif _u in {"gb-hour", "per gb-hour"} and "ram" not in refs:
                refs["ram"] = r
    # Hide the Value column from the deliverable. It stays in the workbook because every
    # pricing formula (Compute/Storage/etc.) references 'Rate Card'!$C$<row>; deleting it
    # would break the math, so it's hidden rather than removed.
    ws.column_dimensions["C"].hidden = True
    return refs


def embed_architecture(ws_po, png_path, anchor_spec=None):
    """Drop this BOM's generated diagram into the Pricing Overview's architecture slot.

    The template no longer carries an architecture picture - only its ANCHOR (the cell
    footprint the picture occupied), kept in the spec as `architecture_anchor`. The image
    itself was the source workbook's own architecture drawing: 637 KB of another customer's
    diagram, decoded and inserted on every export just to be overwritten. Storing the
    footprint and nothing else means an export with no diagram simply has no picture, and
    one with a diagram gets ours in exactly the right place.
    """
    if not png_path or not Path(png_path).exists() or not anchor_spec:
        return False
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU

    fr = anchor_spec["from"]
    m1 = AnchorMarker(col=fr["col"], colOff=fr["colOff"], row=fr["row"], rowOff=fr["rowOff"])
    new = XLImage(str(png_path))
    # Anchor aspect-preserving so the diagram isn't stretched/compressed. A TwoCellAnchor
    # forces the image to fill the exact cell rectangle (whatever its shape), which squishes
    # a tall/portrait diagram. Instead, fit the WIDTH to the anchor's column span and derive
    # the HEIGHT from the image's own proportions via a OneCellAnchor.
    img_w = float(getattr(new, "width", 0) or 0)
    img_h = float(getattr(new, "height", 0) or 0)
    to = anchor_spec.get("to")
    box_w_px = 0
    if to and img_w > 0:
        for c in range(fr["col"], to["col"]):          # columns spanned (0-indexed)
            cd = ws_po.column_dimensions.get(get_column_letter(c + 1))
            w = cd.width if (cd and cd.width) else 8.43
            box_w_px += int(round(w * 7 + 5))          # width chars -> pixels (default font)
    disp_h = 0
    if box_w_px > 0 and img_w > 0:
        disp_w = box_w_px
        disp_h = int(round(box_w_px * img_h / img_w))  # keep the native aspect ratio
        new.anchor = OneCellAnchor(_from=m1, ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(disp_w), cy=pixels_to_EMU(disp_h)))
    elif anchor_spec.get("type") == "TwoCellAnchor" and to:
        m2 = AnchorMarker(col=to["col"], colOff=to["colOff"], row=to["row"], rowOff=to["rowOff"])
        new.anchor = TwoCellAnchor(_from=m1, to=m2)
    else:
        ext = anchor_spec.get("ext_emu")
        new.anchor = OneCellAnchor(
            _from=m1, ext=XDRPositiveSize2D(cx=ext["cx"], cy=ext["cy"]) if ext else None)
    ws_po.add_image(new)

    # Row (1-indexed) just past the picture's bottom edge, so the caller can place footnotes
    # below the diagram no matter how tall the DR section made it. Walk rows from the anchor
    # top, summing their pixel heights until we've covered the image height.
    start_row0 = fr["row"]                          # 0-indexed anchor top
    if disp_h > 0:
        acc = -(fr.get("rowOff", 0) / 9525.0)       # EMU -> px of the top offset
        r0 = start_row0
        while acc < disp_h and r0 < start_row0 + 400:
            rd = ws_po.row_dimensions.get(r0 + 1)
            h_pt = rd.height if (rd and rd.height) else 15.0
            acc += h_pt * 96.0 / 72.0                # points -> pixels
            r0 += 1
        bottom_row = r0 + 1                          # 1-indexed, one row below the image
    else:
        bottom_row = start_row0 + 44                 # fallback: old ~43-row footprint
    return bottom_row


CUSTOMER_TOKEN = "{{CUSTOMER}}"


def _apply_customer_name(wb, bom_name):
    """Swap the template's {{CUSTOMER}} token for the BOM name the user typed.

    The template is deliberately customer-neutral - it must never ship one client's name
    (or servers, or numbers) inside another client's deliverable. Every customer-specific
    string in the spec is the token; this is the only place a real name enters the
    workbook. With no BOM name, it degrades to a generic "the customer".
    """
    name = (bom_name or "").strip()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and CUSTOMER_TOKEN in v:
                    if not name:
                        # Mid-sentence the generic reads "the customer"; as a title it
                        # would read "The customer OCI Migration", so trim the token out.
                        v = (v.replace(CUSTOMER_TOKEN + " ", "the customer ")
                             if not v.startswith(CUSTOMER_TOKEN)
                             else v[len(CUSTOMER_TOKEN):].lstrip())
                        v = v.replace(CUSTOMER_TOKEN, "the customer")
                    else:
                        v = v.replace(CUSTOMER_TOKEN, name)
                    cell.value = v


def _set_toc(ws, bom_name):
    name = (bom_name or "").strip()
    title = name or "OCI Bill of Materials"
    ws["A1"] = f"{title} - OCI BOM"
    ws["B5"] = title
    ws["B6"] = f"{title} OCI estimate"
    ws["B7"] = "OCI BOM + Architecture Generator"
    ws["A8"] = "Generated On"
    ws["B8"] = datetime.date.today().isoformat()
    ws.row_dimensions[8].hidden = False
    ws.row_dimensions[8].height = 20
    ws["A10"] = "Workbook Status"
    ws["B10"] = "Generated from current estimate"
    ws["B11"] = (
        "Includes only workloads, services, pricing selections, and ramp inputs present "
        "in the current estimator session."
    )
    ws["B12"] = "Oracle Cloud Infrastructure"


def _set_assumptions(ws, servers, shape_label="", hours=None):
    """Replace the reference customer's narrative with facts from this export."""
    compute_count = len(servers or [])
    ws["A5"] = (
        f"• This export contains {compute_count:,} priced compute workload "
        f"{'row' if compute_count == 1 else 'rows'} from the current estimate."
    )
    ws["A6"] = (
        "• Service, application, storage, and architecture details are populated only "
        "when they exist in the current estimator state."
    )
    ws["A7"] = (
        "• No workload names, capacities, migration waves, site counts, or scope "
        "assumptions are inherited from the styling reference."
    )
    ws["A8"] = (
        "• Pricing remains a planning estimate and should be validated against the "
        "current Oracle ordering documents."
    )
    assumptions = [
        ("Scope", "Only current estimator workloads and selected OCI services are included."),
        ("Pricing catalog", "SKU rates are copied from the application's current OCI catalog at export time."),
        ("Selected compute shape", shape_label or "No compute shape was selected."),
        ("Monthly usage", (
            f"The default is {float(hours):g} hours per month; workload-specific values "
            "from the estimate take precedence." if hours else
            "Monthly usage is carried from each current estimate row."
        )),
        ("Licensing", "Licensing is included only when selected or detected in the current estimate."),
        ("Optional services", "Unselected services contribute no quantity or cost."),
        ("Reference workbook", "Used for visual styling only; its customer data and assumptions are excluded."),
    ]
    for row, (label, value) in enumerate(assumptions, 11):
        ws[f"A{row}"] = label
        ws[f"C{row}"] = value


# One-line purpose / primary-use blurbs for the Table of Contents, keyed by sheet name.
_TOC_DESC = {
    "Assumptions": ("Workbook methodology and scope assumptions.",
                    "Audit how the BOM was assembled and priced."),
    "Rate Card": ("SKU unit rates used throughout the workbook.",
                  "Review compute, storage, networking, and service rates."),
    "Pricing Overview": ("Cost baseline, OCI discount, and the source-vs-OCI comparison.",
                         "Start here for total monthly cost and savings."),
    "Service Mapping": ("Per-line source-to-OCI mapping, grouped by product, with the net OCI total.",
                        "See how each source service maps and prices; the total ties to the Pricing Overview."),
    "Compute": ("VM inventory with source-to-OCI shape mapping and compute pricing.",
                "Review each server's shape mapping and adjust monthly hours."),
    "Storage": ("Storage services populated by the current estimate.",
                "Review mapped storage quantities, rates, and cost."),
    "Networking": ("Networking services populated by the current estimate.",
                   "Review mapped networking quantities, rates, and cost."),
    "DR": ("Disaster-recovery services populated by the current estimate.",
           "Review mapped DR quantities, rates, and cost."),
    "Security KMS": ("Security services populated by the current estimate.",
                     "Review mapped security quantities, rates, and cost."),
    "Consumption Ramp": ("Month-by-month consumption ramp and 5-year projection.",
                         "See the configured ramp and cumulative cost."),
    "Annexure Addendum to Storage": ("Optional storage detail from the current estimate.",
                                     "Review only when current storage inputs populate it."),
    "Applications Migrated to OCI": ("Application-level migration summary.",
                                     "Filter by application for VM counts, tier mix, and cost."),
    "Database": ("Database service mapping and pricing detail.",
                 "Review DB shapes, editions, and storage."),
    "Integration": ("Integration service SKU mapping and pricing detail.",
                    "Review OIC, API, event, and messaging services."),
    "Observability": ("Observability and management service detail.",
                      "Review logging, monitoring, and management usage."),
    "AI and Machine Learning": ("AI and machine-learning service detail.",
                                "Review GPU, AI, and model-service mappings."),
    "Licensing": ("Third-party licensing SKU detail.",
                  "Review license quantities and non-discountable cost."),
    "Other Services": ("Additional OCI service mapping and pricing detail.",
                       "Review services outside the core infrastructure tabs."),
    "Notes + Assumptions": ("Per-service mapping notes and assumptions.",
                            "Reference for how each source service was mapped."),
}


def _rebuild_toc(wb, bom_name=""):
    """Rewrite the Table of Contents table (rows 16+) to list EVERY visible sheet in the
    workbook, each hyperlinked to its tab. The template shipped a stale, misaligned list;
    this regenerates it after all sheets are added/removed/hidden."""
    from copy import copy
    from openpyxl.worksheet.hyperlink import Hyperlink
    if "Table of Contents" not in wb.sheetnames:
        return
    ws = wb["Table of Contents"]
    FIRST = 16
    sa, sb, sc = (copy(ws.cell(FIRST, c)._style) for c in (1, 2, 3))
    link_font = copy(ws.cell(FIRST, 1).font)   # blue underlined hyperlink style
    # Clear the old (stale) table body.
    for r in range(FIRST, FIRST + 60):
        for c in (1, 2, 3):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None
    skip = {"Table of Contents", "_workflow"}
    r = FIRST
    for name in wb.sheetnames:
        if name in skip:
            continue
        if getattr(wb[name], "sheet_state", "visible") != "visible":
            continue
        purpose, use = _TOC_DESC.get(
            name, (f"{name} product-group detail.", f"Line-item detail for the {name} group."))
        a = ws.cell(r, 1, name)
        a._style = copy(sa)
        a.font = copy(link_font)
        a.hyperlink = Hyperlink(ref=a.coordinate, location=f"'{name}'!A1", display=name)
        b = ws.cell(r, 2, purpose); b._style = copy(sb)
        c = ws.cell(r, 3, use); c._style = copy(sc)
        r += 1
    # Blank any trailing template rows below the new list (already cleared above).
    return r


def _clear_cell_note(ws, coord):
    """Strip the hidden tooltip on a cell: a hover comment/note AND any data-validation
    input prompt (Excel shows a DV prompt as a note when the cell is selected). B9 shipped
    a stale prompt titled 'E6 Optimization' that no longer describes what the cell does."""
    try:
        ws[coord].comment = None
    except Exception:
        pass
    try:
        dvs = list(ws.data_validations.dataValidation)
    except Exception:
        return
    keep = []
    for dv in dvs:
        sq = str(dv.sqref).strip()
        if coord in sq:
            if sq == coord:
                continue                    # covers only this cell -> drop entirely
            dv.showInputMessage = False     # covers more -> just silence the prompt
            dv.prompt = None
            dv.promptTitle = None
        keep.append(dv)
    ws.data_validations.dataValidation = keep


# Cloud-bill service groups -> Pricing Overview line (same grouping the Product Breakdown
# uses). Compute stays in the Compute sheet (lines 13-15). Everything else rolls into the
# matching Overview line so the template total includes the whole bill, not just compute.
_CLOUD_GROUP_TO_OVERVIEW_ROW = {
    "Storage": 16, "Networking": 18, "Security": 19,
    "Database": 16, "Obs. & Management": 16, "Other Services": 16,
    "AI & Machine Learning": 16, "DevOps": 16, "Marketplace": 16, "Support": 16,
}


def _source_shape_label(pr):
    """The source instance/shape a compute row was mapped FROM (e.g. AWS m6a.4xlarge,
    Azure E8s v5). Prefers the detected source instance, then the AWS usage-type suffix
    (BoxUsage:<type>), then a clean source product family."""
    sce = pr.get("sourceCloudEstimate") or {}
    if sce.get("instance"):
        return str(sce["instance"])
    ut = str(pr.get("sourceUsageType") or "")
    if ":" in ut:
        cand = ut.split(":")[-1].strip()
        if cand and not cand.replace(".", "").replace(",", "").isdigit():
            return cand
    sp = str(pr.get("sourceProduct") or "").strip()
    return sp if (sp and "$" not in sp and "per " not in sp.lower()) else ""


def _cloud_effective_hours(pr):
    """Effective monthly hours for a cloud-bill compute row = billed OCPU-hours / OCPUs.
    The bill meters actual usage, so a line can be far more (or less) than 730 hours; using
    the effective hours makes the Compute sheet's OCPU x hours x rate reproduce the app's
    actual cost. Falls back to the row's hoursPerMonth when the OCPU-hr quantity is absent."""
    specs = pr.get("specs") or {}
    ocpus = float(specs.get("ocpus") or 0)
    if ocpus > 0:
        for li in (pr.get("lineItems") or []):
            d = (li.get("description") or "").lower()
            if "ocpu" in d and ("hr" in d or "hour" in d):
                q = float(li.get("quantity") or 0)
                if q > 0:
                    return round(q / ocpus, 4)
    return float(pr.get("hoursPerMonth") or 0) or None


def _add_cloud_bill_services(wb, pricing):
    """Set the Pricing Overview lines from the app's EXACT per-row monthly so the whole
    cloud bill ties out - every service, not just compute.

    Each bill line's monthly is allocated to one Overview line by category. Compute rows
    carry non-OCPU/RAM costs too (attached storage, data transfer...), so compute is split
    into its OCPU / RAM / other-compute line-item sums. The Overview lines are written as
    VALUES (replacing the =SUM(Compute!...) formulas) because the Compute sheet re-derives
    from spec x hours and can't reproduce a metered bill's true totals. Sum of lines 13-20
    then equals the app's OCI monthly exactly; 3rd-party licensing stays on line 21."""
    import bom_export
    ws = wb["Pricing Overview"]
    comp_ocpu = comp_ram = comp_other = 0.0
    svc = {}                       # overview row -> total
    other_non_storage = False
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        grp = bom_export._cloud_product_group(r.get("ociServiceCategory"), r.get("sourceService"))
        if grp == "Compute":
            for li in (r.get("lineItems") or []):
                d = (li.get("description") or "").lower()
                m = float(li.get("monthly") or 0)
                if "ocpu" in d and ("hr" in d or "hour" in d):
                    comp_ocpu += m
                elif "memory" in d and ("hr" in d or "hour" in d):
                    comp_ram += m
                else:
                    comp_other += m
        else:
            row = _CLOUD_GROUP_TO_OVERVIEW_ROW.get(grp, 16)
            svc[row] = svc.get(row, 0.0) + float(r.get("monthly") or 0)
            if row == 16 and grp != "Storage":
                other_non_storage = True

    # Compute lines: OCPU -> 13, RAM -> 14, everything else on compute rows -> 15.
    ws["B13"] = round(comp_ocpu, 2)
    ws["B14"] = round(comp_ram, 2)
    ws["B15"] = round(comp_other, 2)
    if comp_other:
        ws["A15"] = "VM Block Storage + attached:"
    # Service lines by group.
    for row, amt in svc.items():
        base = ws[f"B{row}"].value
        base = 0.0 if isinstance(base, str) else float(base or 0)
        ws[f"B{row}"] = round(base + amt, 2)
    if 16 in svc and other_non_storage:
        ws["A16"] = "Storage / Other OCI Services:"


# The 11 product groups (matches the Product Breakdown) and the template sheets that
# already cover four of them; the rest get a generated detail sheet.
_PRODUCT_GROUPS = [
    "Compute", "Database", "Storage", "Networking", "Support",
    "Obs. & Management", "Other Services", "Security",
    "AI & Machine Learning", "DevOps", "Marketplace",
]
_GROUP_EXISTING_SHEET = {
    "Compute": "Compute", "Storage": "Storage",
    "Networking": "Networking", "Security": "Security KMS",
}
_GROUP_SHEET_NAME = {
    "Obs. & Management": "Obs. and Management",
    "AI & Machine Learning": "AI and Machine Learning",
}


def _aggregate_product_groups(pricing):
    """Group the priced rows into the 11 product groups, returning
    {group: {"aws", "oci", "items": {(awsService, ociProduct): {"aws","oci"}}}}."""
    import bom_export
    groups = {}
    for r in (pricing or {}).get("rows", []):
        if (r.get("costAction") or "") == "remove":
            continue
        grp = bom_export._cloud_product_group(
            r.get("ociServiceCategory"), r.get("sourceService") or "Other",
            r.get("ociProduct"))
        aws = float(r.get("sourceMonthlyCost") or 0)
        oci = float(r.get("monthly") or 0)
        g = groups.setdefault(grp, {"aws": 0.0, "oci": 0.0, "items": {}})
        g["aws"] += aws
        g["oci"] += oci
        key = (_clean(r.get("sourceService")) or "Other",
               _clean(r.get("ociProduct")) or "Needs review")
        it = g["items"].setdefault(key, {"aws": 0.0, "oci": 0.0})
        it["aws"] += aws
        it["oci"] += oci
    return groups


def _add_product_group_topics(wb, pricing):
    """Cloud-bill: put a 'Cost by Product Group' summary of ALL 11 topics on the Pricing
    Overview, and add a detail sheet for every group that has cost and doesn't already have
    a dedicated sheet (Compute/Storage/Networking/Security KMS already exist)."""
    import bom_export
    groups = _aggregate_product_groups(pricing)
    ctr = Alignment(horizontal="center", vertical="center")
    hdr_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    lbl_font = Font(name="Calibri", size=11, bold=True)
    money_fmt = "#,##0.00"

    # ---- per-group detail sheets for groups without a dedicated sheet ----
    for grp in _PRODUCT_GROUPS:
        g = groups.get(grp)
        if not g or (g["oci"] <= 0 and g["aws"] <= 0):
            continue
        if grp in _GROUP_EXISTING_SHEET:
            continue
        name = (_GROUP_SHEET_NAME.get(grp, grp))[:31]
        if name in wb.sheetnames:
            continue
        ws2 = wb.create_sheet(name)
        ws2.column_dimensions["A"].width = 46
        ws2.column_dimensions["B"].width = 46
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18
        ws2["A1"] = f"{grp} - OCI mapped services"
        ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
        for c, txt in ((1, "AWS Service"), (2, "OCI Product"),
                       (3, "AWS Monthly"), (4, "OCI Monthly")):
            cell = ws2.cell(3, c)
            cell.value = txt
            cell.font = hdr_font
            cell.fill = PatternFill("solid", fgColor="FF4472C4")
            cell.alignment = ctr
        rr = 4
        for (aws_svc, oci_prod), v in sorted(g["items"].items(), key=lambda kv: -kv[1]["oci"]):
            ws2.cell(rr, 1).value = aws_svc
            ws2.cell(rr, 2).value = oci_prod
            ws2.cell(rr, 3).value = round(v["aws"], 2)
            ws2.cell(rr, 4).value = round(v["oci"], 2)
            ws2.cell(rr, 3).number_format = money_fmt
            ws2.cell(rr, 4).number_format = money_fmt
            rr += 1
        ws2.cell(rr, 1).value = "Total"
        ws2.cell(rr, 1).font = lbl_font
        ws2.cell(rr, 3).value = f"=SUM(C4:C{rr-1})"
        ws2.cell(rr, 4).value = f"=SUM(D4:D{rr-1})"
        ws2.cell(rr, 3).number_format = money_fmt
        ws2.cell(rr, 4).number_format = money_fmt
        ws2.cell(rr, 3).font = lbl_font
        ws2.cell(rr, 4).font = lbl_font


def _set_optimization(ws_compute, rightsized=False, ocpu_pct=0.0, ram_pct=0.0, is_ax=False):
    """Relabel and fill the Compute-optimization block.

    The optimization does NOT discount the price directly - it shrinks each VM's OCPU and
    RAM quantities (columns J/K) via ceil(value*(1-pct)), floored at 2, and the price then
    flows from the smaller sizing. B9 is a record of what was applied ("% approximation"),
    not a live multiplier. For Ax shapes the base %% deepens with the source instance's
    generation gap (x2 per generation behind the first), so each row can differ.
    """
    ws_compute["A8"] = "Compute optimization"
    ws_compute["A9"] = "% approximation"
    if rightsized and (ocpu_pct or ram_pct):
        lo, hi = sorted({int(round(ocpu_pct * 100)), int(round(ram_pct * 100))})
        ws_compute["B9"] = f"{lo}–{hi}%" if lo != hi else f"{hi}%"
        if is_ax:
            ws_compute["C9"] = (
                f"% approximation - Ax base OCPU ~{int(round(ocpu_pct*100))}%, RAM ~"
                f"{int(round(ram_pct*100))}%, doubled (×2) for each generation the source "
                "instance is behind the newest OCI generation (1 gen = base, 2 gens = ×2, "
                "3 gens = ×4, …), capped at 95%. Applied to the sizing in columns J and K "
                "as ceil(value × (1 − %)) with a floor of 2, so each row can differ. Price "
                "follows the reduced sizing; this % is not a direct price discount."
            )
        else:
            ws_compute["C9"] = (
                f"% approximation - OCPU reduced ~{int(round(ocpu_pct*100))}%, RAM ~"
                f"{int(round(ram_pct*100))}%, applied to the sizing in columns J and K as "
                "ceil(value × (1 − %)) with a floor of 2. Price follows the reduced "
                "sizing; this % is not a direct price discount."
            )
    else:
        ws_compute["B9"] = "0%"
        ws_compute["C9"] = (
            "% approximation - no compute optimization applied (Rightsize off, or the "
            "selected shape is not eligible). Sizing is carried over as-is."
        )
    ws_compute["F9"] = (
        "Rightsizing is reflected only when it is enabled in the current estimate."
    )
    # Strip any hidden note/comment or data-validation input prompt the template attached
    # to B9 (it shows as a tooltip on the cell and is separate from the visible C9 text).
    _clear_cell_note(ws_compute, "B9")


MONTH_COLS = "BCDEFGHIJKLM"  # Consumption Ramp grid: months 1..12


def _add_licensing_line(ws_po, windows_monthly):
    """Add a '3rd Party Licensing (Windows)' line to the Pricing Overview baseline and
    re-total. Done by rewriting cells (not insert_rows) so merges / conditional formats /
    images stay intact:  B21 = licensing, B22 = Total Monthly = SUM(B13:B21), B23 = Annual."""
    from copy import copy
    for col in ("A", "B"):
        ws_po[f"{col}23"]._style = copy(ws_po[f"{col}22"]._style)   # annual-total style
        ws_po[f"{col}22"]._style = copy(ws_po[f"{col}21"]._style)   # monthly-total style
        ws_po[f"{col}21"]._style = copy(ws_po[f"{col}20"]._style)   # ordinary line style
    ws_po["A21"] = "3rd Party Licensing (Windows):"
    ws_po["B21"] = round(float(windows_monthly or 0), 2)
    ws_po["A22"] = "Total Monthly Cost:"
    ws_po["B22"] = "=SUM(B13:B21)"
    ws_po["A23"] = "Total Annual Cost:"
    ws_po["B23"] = "=B22*12"


def _add_extra_services(wb, extra_priced):
    """Fold app-added OCI services into the Pricing Overview.

    Each service rolls into the category line it belongs to (Networking -> B18,
    Security -> B19, everything else -> B16), by appending its monthly cost to that line's
    formula. Those lines all sit inside the Total's SUM(B13:B21), so the workbook total
    picks the extras up automatically and still ties out. The added services are also
    itemized on their matching service tabs for traceability.

    `extra_priced` is the authoritative list from oci_catalog.price_extras().
    """
    import oci_catalog
    if not extra_priced:
        return
    ws = wb["Pricing Overview"]
    sums = {}
    non_storage_into_16 = False
    for s in extra_priced:
        category = _service_category(s.get("group"), s.get("name"), s.get("name"))
        row = 20 if category == "DR" else oci_catalog.GROUP_TO_OVERVIEW_ROW.get(
            s["group"], 16)
        sums[row] = sums.get(row, 0.0) + float(s["monthly"] or 0)
        if row == 16 and s["group"] not in ("Storage", "Database"):
            non_storage_into_16 = True

    for row, amount in sums.items():
        cell = ws[f"B{row}"]
        base = cell.value
        add = round(amount, 2)
        if isinstance(base, str) and base.startswith("="):
            cell.value = f"{base}+{add}"
        else:
            cell.value = round(float(base or 0) + add, 2)

    if 16 in sums and non_storage_into_16:
        ws["A16"] = "Storage / Other OCI Services:"

_SERVICE_SHEETS = {
    "Compute": "Compute",
    "Storage": "Storage",
    "Networking": "Networking",
    "Database": "Database",
    "Integration": "Integration",
    "Security": "Security KMS",
    "Observability": "Observability",
    "AI & Machine Learning": "AI and Machine Learning",
    "Licensing": "Licensing",
    "Other Services": "Other Services",
    "DR": "DR",
}


def _service_category(category, product="", source_service=""):
    """Normalize app/catalog and cloud-bill group names to the service tabs used by the BOM."""
    raw = _clean(category)
    key = _norm(raw)
    aliases = {
        "compute": "Compute",
        "storage": "Storage",
        "networking": "Networking",
        "database": "Database",
        "integration": "Integration",
        "devops": "Integration",
        "security": "Security",
        "observability": "Observability",
        "obs management": "Observability",
        "ai machine learning": "AI & Machine Learning",
        "ai and machine learning": "AI & Machine Learning",
        "licensing": "Licensing",
        "other services": "Other Services",
        "support": "Other Services",
        "marketplace": "Other Services",
    }
    combined = _norm(f"{product} {source_service}")
    if "full stack disaster recovery" in combined or "full stack dr" in combined:
        return "DR"
    return aliases.get(key, "Other Services")


def _service_detail_lines(pricing, extra_priced):
    """Return a full, per-SKU paper trail grouped by the app's service categories."""
    import bom_export
    grouped = {}

    def add(category, line):
        grouped.setdefault(category, []).append(line)

    source_cloud = _clean((pricing or {}).get("sourceCloud")).upper()
    for row in (pricing or {}).get("rows", []):
        if (row.get("costAction") or "") == "remove":
            continue
        # VM rows already have their complete compute/storage contract on the Compute
        # sheet. Some on-prem estimates do not set ociServiceCategory, so category text
        # alone cannot be used to recognize them. Keep only third-party license line
        # items from those rows for the Licensing tab; never duplicate the VM's OCPU,
        # RAM, and block SKUs into Other Services.
        if (row.get("specs") or {}).get("ocpus"):
            source = _clean(row.get("sourceService")) or "uploaded inventory"
            for item in row.get("lineItems") or []:
                description = _clean(item.get("description"))
                if not any(term in _norm(description)
                           for term in ("windows", "sql server", "license", "licence")):
                    continue
                note = f"Mapped from {source_cloud + ' ' if source_cloud else ''}{source}."
                mapping = _clean(item.get("mapping"))
                if mapping:
                    note += f" {mapping}"
                add("Licensing", {
                    "component": description or _clean(row.get("ociProduct")),
                    "sku": _clean(item.get("sku")),
                    "unit": _clean(item.get("unit")) or "usage",
                    "rate": float(item.get("rate") or 0),
                    "qty": item.get("quantity"),
                    "hours": item.get("hours"),
                    "monthly": float(item.get("monthly") or 0),
                    "notes": note,
                })
            continue
        raw_cat = row.get("ociServiceCategory")
        if not raw_cat:
            raw_cat = bom_export._cloud_product_group(
                raw_cat, row.get("sourceService"), row.get("ociProduct"))
        category = _service_category(
            raw_cat, row.get("ociProduct"), row.get("sourceService"))
        items = row.get("lineItems") or [{
            "sku": (row.get("fullServiceMapping") or {}).get("sku"),
            "description": row.get("ociProduct"),
            "quantity": 1,
            "unit": "month",
            "rate": row.get("monthly"),
            "monthly": row.get("monthly"),
        }]
        for i, item in enumerate(items):
            source = _clean(row.get("sourceService")) or "uploaded inventory"
            mapping = _clean(item.get("mapping"))
            note = f"Mapped from {source_cloud + ' ' if source_cloud else ''}{source}."
            if mapping:
                note += f" {mapping}"
            add(category, {
                "component": _clean(row.get("ociProduct")) if i == 0 else "",
                "sku": _clean(item.get("sku")),
                "unit": _clean(item.get("unit")) or "usage",
                "rate": float(item.get("rate") or 0),
                "qty": item.get("quantity"),
                "hours": item.get("hours"),
                "monthly": float(item.get("monthly") or 0),
                "notes": note,
            })

    for service in extra_priced or []:
        category = _service_category(
            service.get("group"), service.get("name"), service.get("name"))
        skus = service.get("skus") or [{
            "sku": service.get("sku"), "desc": service.get("name"),
            "qty": service.get("qty"), "rate": service.get("rate"),
            "hours": service.get("hours"), "monthly": service.get("monthly"),
        }]
        for i, item in enumerate(skus):
            note = "Added in the Networking and Other Services step."
            sizing = _clean(service.get("sizing"))
            if sizing:
                note += f" {sizing}"
            add(category, {
                "component": _clean(service.get("name")) if i == 0 else _clean(item.get("desc")),
                "sku": _clean(item.get("sku")),
                "unit": _clean(service.get("unit")) or _clean(service.get("basis")),
                "rate": float(item.get("rate") or 0),
                "qty": item.get("qty"),
                "hours": item.get("hours"),
                "monthly": float(item.get("monthly") or 0),
                "notes": note,
            })
    return grouped


def _add_spec_sheet_images(ws, spec, source_sheet):
    """Copy the reference workbook's presentation-only banner onto a generated sheet."""
    source = next((s for s in spec.get("sheets", []) if s.get("name") == source_sheet), None)
    if not source:
        return
    for im in source.get("images", []):
        data = base64.b64decode(spec["images"][im["image_ref"]]["base64"])
        img = XLImage(io.BytesIO(data))
        anchor = im["anchor"]
        fr = anchor["from"]
        m1 = AnchorMarker(
            col=fr["col"], colOff=fr["colOff"], row=fr["row"], rowOff=fr["rowOff"])
        if anchor["type"] == "TwoCellAnchor" and "to" in anchor:
            to = anchor["to"]
            m2 = AnchorMarker(
                col=to["col"], colOff=to["colOff"], row=to["row"], rowOff=to["rowOff"])
            img.anchor = TwoCellAnchor(_from=m1, to=m2)
        else:
            ext = anchor.get("ext_emu")
            img.anchor = OneCellAnchor(
                _from=m1,
                ext=XDRPositiveSize2D(cx=ext["cx"], cy=ext["cy"]) if ext else None)
        ws.add_image(img)


def _service_style_prototype(wb):
    """Capture the reference Networking styles before that sheet is repopulated."""
    from copy import copy
    src = wb["Networking"]
    return {
        "header": [copy(src.cell(10, c)._style) for c in range(1, 9)],
        "odd": [copy(src.cell(11, c)._style) for c in range(1, 9)],
        "even": [copy(src.cell(12, c)._style) for c in range(1, 9)],
        "summary": {coord: copy(src[coord]._style)
                    for coord in ("A4", "A5", "B5", "C5", "D5", "F4", "F5", "A6", "A9")},
    }


def _write_service_region(ws, sheet_name, lines, styles, start_row=10, replace=False):
    """Populate one template-styled service detail table without feeding BOM totals twice."""
    from copy import copy
    end_clear = max(120, start_row + len(lines) + 4)
    if replace:
        # The reference tabs carried include/exclude dropdowns and conditional formatting
        # for that customer's scenarios. The generated per-SKU table does not use them.
        ws.data_validations.dataValidation = []
        try:
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass
        for merged in list(ws.merged_cells.ranges):
            if merged.min_row >= 4 and merged.max_row <= end_clear:
                ws.unmerge_cells(str(merged))
        clear_cols = max(16, ws.max_column)
        for r in range(4, end_clear + 1):
            for c in range(1, clear_cols + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        ws["A1"] = sheet_name
        ws["A2"] = sheet_name
        ws["A4"] = f"{sheet_name} BOM Summary"
        ws["A4"]._style = copy(styles["summary"]["A4"])
        ws["A5"] = "Mapped Monthly"
        ws["A5"]._style = copy(styles["summary"]["A5"])
        ws["C5"] = "Mapped Annual"
        ws["C5"]._style = copy(styles["summary"]["C5"])
        ws["F4"] = "Mapping Notes / Assumptions"
        ws["F4"]._style = copy(styles["summary"]["F4"])
        ws["F5"] = (
            "These rows are the complete per-SKU mapping from the app. They are a detail "
            "view only; Pricing Overview remains the workbook total."
        )
        ws["F5"]._style = copy(styles["summary"]["F5"])
        ws["A6"] = (
            "Every populated service, source mapping, unit rate, quantity, and monthly "
            "amount is carried from the estimator output."
        )
        ws["A6"]._style = copy(styles["summary"]["A6"])
        ws["A9"] = "Service pricing rows below reconcile to the mapped monthly summary."
        ws["A9"]._style = copy(styles["summary"]["A9"])
        header_row = start_row
    else:
        for r in range(start_row, end_clear + 1):
            for c in range(1, 9):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        ws.cell(start_row, 1).value = f"{sheet_name} - mapped service detail"
        ws.cell(start_row, 1)._style = copy(styles["summary"]["A4"])
        header_row = start_row + 1

    headers = ["Component", "SKU", "Unit", "Unit Rate", "Qty / Input",
               "Hours / Volume", "Est. Monthly", "Source Mapping / Notes"]
    for c, value in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, value)
        cell._style = copy(styles["header"][c - 1])

    first = header_row + 1
    for i, line in enumerate(lines):
        r = first + i
        row_styles = styles["odd"] if i % 2 == 0 else styles["even"]
        values = [
            line.get("component"), line.get("sku"), line.get("unit"),
            round(float(line.get("rate") or 0), 6), line.get("qty"),
            line.get("hours"), round(float(line.get("monthly") or 0), 2),
            line.get("notes"),
        ]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(r, c, value)
            cell._style = copy(row_styles[c - 1])
        ws.cell(r, 4).number_format = "#,##0.000000"
        ws.cell(r, 7).number_format = "#,##0.00"
        ws.row_dimensions[r].hidden = False

    total_row = first + len(lines)
    for c in range(1, 9):
        ws.cell(total_row, c)._style = copy(styles["header"][c - 1])
    ws.cell(total_row, 1).value = f"Total {sheet_name}"
    ws.cell(total_row, 7).value = (
        f"=SUM(G{first}:G{total_row - 1})" if lines else 0)
    ws.cell(total_row, 7).number_format = "#,##0.00"
    ws.row_dimensions[header_row].hidden = False
    ws.row_dimensions[total_row].hidden = False
    if replace:
        ws["B5"] = f"=G{total_row}"
        ws["B5"]._style = copy(styles["summary"]["B5"])
        ws["D5"] = "=B5*12"
        ws["D5"]._style = copy(styles["summary"]["D5"])
        ws.freeze_panes = f"A{first}"
        for r in range(total_row + 1, end_clear + 1):
            ws.row_dimensions[r].hidden = True
    ws.sheet_state = "visible"
    return total_row


def _write_service_tabs(wb, pricing, extra_priced, spec, compute_last, storage_count):
    """Map every non-VM service into the matching Polaris-styled workbook tab."""
    grouped = _service_detail_lines(pricing, extra_priced)
    if not grouped:
        return
    styles = _service_style_prototype(wb)
    source = wb["Networking"]

    # Create missing category tabs from the reference Networking layout before that
    # source sheet is repopulated. copy_worksheet preserves its cells/styles/merges;
    # images are copied explicitly from the build spec.
    for category in grouped:
        name = _SERVICE_SHEETS[category]
        if name in wb.sheetnames:
            continue
        ws = wb.copy_worksheet(source)
        ws.title = name
        _add_spec_sheet_images(ws, spec, "Networking")

    for category, lines in grouped.items():
        name = _SERVICE_SHEETS[category]
        ws = wb[name]
        if name in ("Networking", "Security KMS") or name not in {
                "Compute", "Storage", "DR"}:
            _write_service_region(ws, name, lines, styles, start_row=10, replace=True)
        elif name == "Compute":
            _write_service_region(
                ws, name, lines, styles, start_row=max(compute_last + 3, 20), replace=False)
        elif name == "Storage":
            _write_service_region(
                ws, name, lines, styles,
                start_row=max(STORAGE_FIRST_ROW + storage_count + 2, 24), replace=False)
        elif name == "DR":
            _write_service_region(ws, name, lines, styles, start_row=10, replace=True)

    # Keep service tabs together and in the same order as the app's service selector.
    after = "Compute"
    for name in ("Storage", "Networking", "Database", "Integration", "Security KMS",
                 "Observability", "AI and Machine Learning", "Licensing",
                 "Other Services", "DR"):
        if name in wb.sheetnames:
            _place_after(wb, name, after)
            after = name


def _repoint_ramp_refs(ws_po, months, include_windows=False):
    """The cost-profile's Year-1 columns sum the ramp's consumption %. Re-point those
    ranges at the actual number of ramp months the app is set to. Year 1 uses at most
    the first 12 ramp months from the generated Consumption % column."""
    y1_last = RAMP_FIRST_MONTH_ROW + min(12, max(1, months)) - 1
    rng = f"'Consumption Ramp'!$B${RAMP_FIRST_MONTH_ROW}:$B${y1_last}"
    # Core infrastructure = compute; Windows 3rd-party licensing (B21) is folded in and
    # ramped alongside it only when it's present in the BOM.
    core = "SUM($B$13:$B$15,$B$21)" if include_windows else "SUM($B$13:$B$15)"
    ws_po["E14"] = f"=IFERROR({core}*SUM({rng}),0)"
    ws_po["F14"] = f"=IFERROR({core}*12,0)"
    ws_po["E15"] = f"=IFERROR($B$16*SUM({rng}),0)"
    ws_po["E17"] = f"=IFERROR($B$20*SUM({rng}),0)"


# A cross-sheet reference token (quoted or bare sheet name + '!' + a cell/range). These
# are captured and stashed so the row-shift below never rewrites another sheet's rows
# (Consumption Ramp B12:B23, Compute!$A$14:..., DR!$B$5, Annexure ...).
_CROSS_SHEET_REF = re.compile(
    r"(?:'[^']*'|[A-Za-z_][\w.]*)!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?")
# A bare (same-sheet) A1 reference: optional $ before column and row.
_BARE_REF = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _shift_same_sheet_rows(ws, lo, hi, delta):
    """Re-point every SAME-SHEET row reference in [lo, hi] by -delta across the whole
    sheet, leaving cross-sheet references untouched. Cross-sheet tokens are stashed as
    placeholders first, the remaining bare refs are shifted, then the tokens restored."""
    def rewrite(formula):
        holds = []

        def stash(m):
            holds.append(m.group(0))
            return "\x01%d\x01" % (len(holds) - 1)

        tmp = _CROSS_SHEET_REF.sub(stash, formula)

        def shift(m):
            col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), int(m.group(4))
            if lo <= row <= hi:
                row -= delta
            return f"{col_abs}{col}{row_abs}{row}"

        tmp = _BARE_REF.sub(shift, tmp)
        for i, tok in enumerate(holds):
            tmp = tmp.replace("\x01%d\x01" % i, tok)
        return tmp

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = rewrite(cell.value)


def _apply_overview_discount(ws, discount, sql_compute=0.0, sql_database=0.0):
    """Add an editable OCI-discount input at A7/B7 (post-relayout layout) and rewire the
    Total Monthly Cost (B18) + OCI Cost Profile to apply it live: discountable OCI services
    x (1 - B7), while 3rd-party licensing is NEVER discounted. Microsoft SQL Server licensing
    is bundled inside the service categories (Compute-grouped rows land in B11, Database-grouped
    rows in B12); it's moved into the 3rd-party licensing line (B17, relabeled Windows + SQL) so
    every formula references cells only. Total Annual (B19 = B18*12) and the comparison blocks
    reference B18/B19, so editing B7 (or any input cell) re-flows the whole sheet."""
    from openpyxl.styles import Font as _F, PatternFill as _PF, Border as _Bd, Side as _Sd, Alignment as _Al
    ws["A7"] = "OCI Discount:"
    ws["A7"].font = _F(bold=True)
    ws["A7"].alignment = _Al(horizontal="right", vertical="center")
    b7 = ws["B7"]
    b7.value = round(max(0.0, min(1.0, float(discount or 0))), 4)
    b7.number_format = "0.0%"
    b7.alignment = _Al(horizontal="center", vertical="center")
    b7.font = _F(bold=True)
    b7.fill = _PF("solid", fgColor="FFF2CC")
    _thin = _Sd(style="thin", color="BF8F00")
    b7.border = _Bd(left=_thin, right=_thin, top=_thin, bottom=_thin)

    sqlc = round(float(sql_compute or 0), 2)     # SQL license bundled in Compute -> B11
    sqld = round(float(sql_database or 0), 2)     # SQL license bundled in Database -> B12

    # Give SQL Server licensing a real home instead of a magic number: move it OUT of the
    # discountable service cells (Compute SQL from B11, Database SQL from B12) and fold it into
    # the 3rd-party licensing line (B17), relabeled "Windows + SQL". Every discount formula can
    # then reference cells only - no embedded constants - and B17 (Windows + SQL) is never
    # discounted. Only what's actually removed is folded in, so nothing is double-counted.
    moved = 0.0
    for ref, amt in (("B11", sqlc), ("B12", sqld)):
        cur = ws[ref].value
        if amt and isinstance(cur, (int, float)):
            ws[ref] = round(float(cur) - amt, 2)
            moved = round(moved + amt, 2)
    if moved:
        cur17 = ws["B17"].value
        ws["B17"] = round((float(cur17) if isinstance(cur17, (int, float)) else 0.0) + moved, 2)
        ws["A17"] = "3rd Party Licensing (Windows + SQL):"

    # Total Monthly: discountable services (SQL now excluded, living in B17) x (1-B7) + B17.
    ws["B18"] = "=SUM(B9:B16)*(1-$B$7)+B17"
    ws["B19"] = "=B18*12"

    # OCI Cost Profile (D:H, rows 10-13): mirror the same math - discount the service refs,
    # leave the licensing ref ($B$17) at list. Cell references only, substituted with
    # _local_ref_sub so a cross-sheet range is never touched.
    subs = [
        ("SUM($B$9:$B$11,$B$17)", "(SUM($B$9:$B$11)*(1-$B$7)+$B$17)"),  # Core infra (B17 = Windows+SQL, at list)
        ("SUM($B$14:$B$15)", "(SUM($B$14:$B$15)*(1-$B$7))"),           # Network, security & KMS
        ("$B$12", "($B$12*(1-$B$7))"),                                  # Object/file storage
        ("$B$16", "($B$16*(1-$B$7))"),                                  # Disaster recovery
    ]
    for r in range(10, 14):
        for c in range(5, 8):                 # E, F, G
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                for old, new in subs:
                    if _local_ref_pattern(old).search(v):
                        cell.value = _local_ref_sub(v, old, new)
                        break                  # one discountable ref type per cell


@lru_cache(maxsize=None)
def _local_ref_pattern(ref):
    """Match `ref` only where it addresses THIS sheet.

    The discount rewrite substitutes cell references like `$B$12` into the Pricing Overview
    formulas. But `$B$12` also occurs inside `'Consumption Ramp'!$B$12:$B$23`, and a plain
    str.replace() rewrote that range into `'Consumption Ramp'!($B$12*(1-$B$7)):$B$23` - not a
    range, and not a formula Excel accepts. Excel refused the workbook with "we found a problem
    with some content"; openpyxl never parses formulas and LibreOffice repaired it silently, so
    only Excel ever objected.

    Anything directly after a `!` belongs to another sheet, so it is left alone. The trailing
    guard stops `$B$12` matching inside `$B$120`.
    """
    return re.compile(r"(?<!!)" + re.escape(ref) + r"(?!\d)")


def _local_ref_sub(formula, ref, replacement):
    """Substitute a same-sheet reference, leaving cross-sheet ranges untouched."""
    return _local_ref_pattern(ref).sub(lambda _m: replacement, formula)


def _relayout_pricing_overview(ws, delta=4):
    """Close the empty gap above the Pricing Baseline so the comparison blocks and the
    architecture diagram can stack beneath it. The Baseline + OCI Cost Profile band
    (rows 12-23) shifts UP by `delta` (-> rows 8-19: Total Monthly -> B18, Total Annual
    -> B19). Same-sheet formula row refs in the band are re-pointed; cross-sheet refs
    (Consumption Ramp, DR, Annexure, Compute) are preserved. The floating notes/title
    that would otherwise collide with the comparison blocks are re-homed below the
    diagram (which the caller repositions to ~row 52)."""
    from copy import copy as _copy
    top, bot = 12, 23                          # source band

    # --- capture the floating notes/title BEFORE the shift, then clear them out of the
    #     band/comparison zone so the shift lands on clean cells ---
    invest_note = ws["D23"].value              # yellow "Oracle is willing to invest..." (D23:H25)
    invest_style = _copy(ws["D23"]._style)
    open_note = ws["B25"].value                # "Open cost items not included above..."
    open_style = _copy(ws["B25"]._style)
    title_note = ws["D27"].value               # "Notional Architecture" (D27:P27)
    title_style = _copy(ws["D27"]._style)
    existing = {str(m) for m in ws.merged_cells.ranges}
    for rng in ("D23:H25", "D27:P27"):
        if rng in existing:
            ws.unmerge_cells(rng)
    for coord in ("D23", "B25", "D27"):
        ws[coord].value = None
        ws[coord].style = "Normal"

    # --- (a) shift the band up. Ascending order so each source row is relocated before a
    #     lower source row overwrites it. Copy value + style + number format per cell. ---
    ncols = ws.max_column
    inside = [m for m in list(ws.merged_cells.ranges)
              if m.min_row >= top and m.max_row <= bot]     # e.g. A12:B12 "Pricing Baseline"
    for m in inside:
        ws.unmerge_cells(str(m))
    for r in range(top, bot + 1):
        for c in range(1, ncols + 1):
            src = ws.cell(r, c)
            dst = ws.cell(r - delta, c)
            dst.value = src.value
            dst._style = _copy(src._style)
            dst.number_format = src.number_format
    for m in inside:
        ws.merge_cells(start_row=m.min_row - delta, start_column=m.min_col,
                       end_row=m.max_row - delta, end_column=m.max_col)
    # null out the now-vacated tail of the old band (rows 20-23)
    for r in range(bot - delta + 1, bot + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.style = "Normal"
    # carry row heights 12-23 -> 8-19, then clear the vacated 20-23
    for r in range(top, bot + 1):
        h = ws.row_dimensions[r].height
        if h is not None:
            ws.row_dimensions[r - delta].height = h
    for r in range(bot - delta + 1, bot + 1):
        ws.row_dimensions[r].height = None

    # --- (b) re-point same-sheet row refs in [12,23] by -delta everywhere on the sheet ---
    _shift_same_sheet_rows(ws, top, bot, delta)

    # --- (c) re-home the title above the diagram; the floating notes are returned to the
    #     caller so they can be dropped BELOW the diagram once its true height is known (the
    #     DR section makes the picture taller, so a fixed note row would collide with it). ---
    if title_note:
        tr = 57                                    # just above the diagram (below the taller comparison block)
        ws[f"D{tr}"] = title_note
        ws[f"D{tr}"]._style = title_style
        ws.merge_cells(start_row=tr, start_column=4, end_row=tr, end_column=16)
    return {
        "open": (open_note, open_style) if open_note else None,
        "invest": (invest_note, invest_style) if invest_note else None,
    }


def _place_overview_notes(ws, start_row, notes):
    """Drop the re-homed Pricing Overview footnotes (open cost items, invest note) at
    `start_row`, each merged A:K and wrapped. Placed below the architecture diagram."""
    r = start_row
    for key in ("open", "invest"):
        payload = (notes or {}).get(key)
        if not payload:
            continue
        value, style = payload
        ws[f"A{r}"] = value
        ws[f"A{r}"]._style = style
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        r += 2


def _zero_unmodeled_sheets(wb):
    """Reset inherited scenario sheets to clean, current-data-only service tables."""
    styles = _service_style_prototype(wb)
    for name in ("Networking", "Security KMS", "DR"):
        ws = wb[name]
        _write_service_region(ws, name, [], styles, start_row=10, replace=True)
        # Preserve the original alternating detail-row styles as the prototype for any
        # current services written later in this build.
        from copy import copy
        for c in range(1, 9):
            ws.cell(11, c)._style = copy(styles["odd"][c - 1])
            ws.cell(12, c)._style = copy(styles["even"][c - 1])

    anx = wb["Annexure Addendum to Storage"]
    anx.data_validations.dataValidation = []
    try:
        anx.conditional_formatting._cf_rules.clear()
    except Exception:
        pass
    for r in (6, 7, 8, 12, 13, 17, 18):
        for c in range(1, 12):
            cell = anx.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for r, label in (
            (8, "Backup Storage Total"),
            (13, "File Storage Total"),
            (18, "Enterprise Storage Total")):
        anx.cell(r, 1).value = label
        anx.cell(r, 8).value = 0
        anx.cell(r, 9).value = 0


# Sheets that always stay visible even when "empty" - they carry the deliverable's
# framing, rates, totals and ramp, so they're meaningful regardless of inventory.
_CORE_SHEETS = {
    "Table of Contents", "Assumptions", "Rate Card", "Pricing Overview",
    "Compute", "Consumption Ramp",
}


def _sheet_has_values(ws, rows, cols):
    """True if any cell in the given rows/cols holds a non-blank, non-zero value."""
    for r in rows:
        for c in cols:
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s and s not in ("0", "0.0", "$0", "$0.00", "-"):
                return True
    return False


def _shape_used_label(shape_used):
    """A readable shape name from a row's shapeUsed (dict payload or plain string)."""
    if isinstance(shape_used, dict):
        return _clean(shape_used.get("shortLabel") or shape_used.get("label")
                      or shape_used.get("key"))
    return _clean(shape_used)


def _place_after(wb, name, after):
    """Move sheet `name` to sit immediately after sheet `after` (no-op if either is missing)."""
    if name not in wb.sheetnames or after not in wb.sheetnames or name == after:
        return
    s = wb[name]
    wb._sheets.remove(s)
    idx = wb.sheetnames.index(after) + 1
    wb._sheets.insert(idx, s)


def _hide_empty_sheets(wb, apps, storage_rows):
    """Hide printout sheets that ended up with no data so the deliverable doesn't ship
    empty sections. Core sheets always stay visible; the rest are hidden (not
    veryHidden, so a user can unhide) when their data region is blank. Hidden sheets
    still compute, so the Pricing Overview references into them are unaffected."""
    def hide(name):
        if name in wb.sheetnames and name not in _CORE_SHEETS:
            wb[name].sheet_state = "hidden"

    if not apps:
        hide("Applications Migrated to OCI")
    if not storage_rows:
        hide("Storage")
    if "Networking" in wb.sheetnames and not _sheet_has_values(
            wb["Networking"], range(12, 140), range(1, 9)):
        hide("Networking")
    if "Security KMS" in wb.sheetnames and not _sheet_has_values(
            wb["Security KMS"], range(12, 140), range(1, 10)):
        hide("Security KMS")
    if "DR" in wb.sheetnames:
        dr = wb["DR"]
        if not _sheet_has_values(dr, range(12, 140), range(1, 9)):
            hide("DR")
    if "Annexure Addendum to Storage" in wb.sheetnames:
        if wb["Annexure Addendum to Storage"]["D6"].value in (None, "", 0):
            hide("Annexure Addendum to Storage")
    # Never leave the active sheet hidden.
    if wb.active is not None and wb.active.sheet_state != "visible":
        for ws in wb.worksheets:
            if ws.sheet_state == "visible":
                wb.active = wb.index(ws)
                break


RAMP_FIRST_MONTH_ROW = 12
RAMP_TPL_LAST_MONTH_ROW = 23      # the reference deliverable ships 12 months
RAMP_TPL_GRID_TITLE_ROW = 27      # "Detailed Consumption Ramp" block starts here
RAMP_TPL_GRID_LAST_ROW = 38
RAMP_MAX_MONTHS = 60

# Component grid rows (relative to the template) -> the Pricing Overview line they scale.
# The consumption ramp scales the OCI SERVICES only (B13:B20), matching the app's ramp,
# whose ceiling is the OCI-services monthly total (pricing.totals.monthly) and excludes
# Windows 3rd-party licensing. Windows (B21) is a separate flat license line, not ramped.
# Each ramp component pulls its steady-state cost from the Pricing Overview baseline band.
# Rows here are the ON-PREM layout positions. In cloud-bill mode the baseline band is shifted
# UP by po_delta rows (see _relayout_pricing_overview), so every row is reduced by po_delta.
# `rows` can list two cells (Networking + Security/KMS are one ramp line). `discountable` marks
# the lines the OCI discount (Pricing Overview $B$7) applies to - everything except 3rd-party
# licensing - so the grid ties to Total Monthly = SUM(discountable)*(1-discount) + licensing.
_GRID_COMPONENTS = [
    ("Compute (OCPUs)", [13], True),
    ("RAM", [14], True),
    ("VM Block Storage", [15], True),
    ("Storage (Object/File)", [16], True),
    ("Storage Backups", [17], True),
    ("Networking + Security / KMS", [18, 19], True),
    ("Disaster Recovery", [20], True),
]


def _ramp_percentages(ramp):
    """The app's ramp as monthly consumption fractions (month spend / steady state).
    Length follows the app's ramp-months toggle."""
    if not isinstance(ramp, dict):
        return None
    ceiling = float(ramp.get("ceiling") or 0)
    monthly = [float(x) for x in (ramp.get("monthly") or [])]
    if ceiling <= 0 or not monthly:
        return None
    return [max(0.0, min(1.0, m / ceiling)) for m in monthly[:RAMP_MAX_MONTHS]]


def _populate_ramp(ws, ramp, include_windows=False, po_delta=0, discount_cell=None):
    """Rebuild the Consumption Ramp for EXACTLY the number of months the app's ramp
    toggle is set to, driven by the app's curve. When Windows 3rd-party licensing is
    present (include_windows), it's added as a ramped component so both ramps carry it.

    Every number is derived from the Pricing Overview baseline so the ramp ties out to it:
      - po_delta shifts the component refs to match the cloud-bill relayout (band moved up 4).
      - discount_cell (Pricing Overview $B$7) is applied to the discountable lines so the grid
        total equals Total Monthly = SUM(discountable)*(1-discount) + licensing.

    Returns the month count so the Pricing Overview's ramp references can be re-pointed.
    """
    from copy import copy
    from openpyxl.utils import get_column_letter

    # Windows licensing is ramped as its own component only when it's actually in the BOM.
    # (Not discountable - the OCI discount never applies to 3rd-party licensing.)
    components = list(_GRID_COMPONENTS)
    if include_windows:
        components.append(("3rd Party Licensing", [21], False))

    pcts = _ramp_percentages(ramp)
    n = len(pcts) if pcts else 12
    n = max(1, min(n, RAMP_MAX_MONTHS))

    # The area carries merges; drop them so every cell is writable while we rebuild.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= RAMP_FIRST_MONTH_ROW:
            ws.unmerge_cells(str(rng))

    def sty(r, c):
        return copy(ws.cell(r, c)._style)

    # ---- capture prototypes from the reference layout BEFORE clearing ----
    # Reuse the reference styles, but collapse the inherited eight-column migration-wave
    # table to four current-data columns: Month, Consumption %, Monthly, Cumulative.
    month_sty = {
        1: sty(12, 1),
        2: sty(12, 6),
        3: sty(12, 7),
        4: sty(12, 8),
    }
    grid_sty = {r: {c: sty(r, c) for c in (1, 2)} for r in range(27, 39)}
    grid_title = ws.cell(27, 1).value
    chart_hdr = ([ws.cell(50, c).value for c in (1, 2)], {c: sty(50, c) for c in (1, 2)})
    chart_sty = {c: sty(51, c) for c in (1, 2)}

    # ---- clear everything below the month header ----
    # Reset BOTH value and style: the reference layout has dark/blue header fills at
    # fixed rows (e.g. the old sequencing and chart headers). When the month table
    # grows past 12 months the layout shifts, so any fill we don't overwrite would be
    # left behind as an orphan block. Blanking the style first prevents that; every
    # cell we actually use gets its style re-applied below.
    for r in range(RAMP_FIRST_MONTH_ROW, 12 + RAMP_MAX_MONTHS + 60):
        for c in range(1, 2 + RAMP_MAX_MONTHS + 2):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None
                cell.style = "Normal"

    # ---- layout: months -> component grid -> chart source ----
    m_first, m_last = RAMP_FIRST_MONTH_ROW, RAMP_FIRST_MONTH_ROW + n - 1
    g_title = m_last + 4
    g_head, g_pct = g_title + 1, g_title + 2
    g_first = g_pct + 1
    g_last = g_first + len(components) - 1
    g_cum = g_last + 1
    c_head = g_cum + 3
    total_col = 2 + n

    # months
    for i in range(n):
        r = m_first + i
        cl = get_column_letter(2 + i)
        for c in range(1, 5):
            ws.cell(r, c)._style = copy(month_sty[c])
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = round(pcts[i], 6) if pcts else None
        ws.cell(r, 3).value = f"=SUM({cl}${g_first}:{cl}${g_last})"
        ws.cell(r, 4).value = f"=SUM($C${m_first}:C{r})"

    # component grid
    ws.cell(g_title, 1).value = grid_title
    ws.cell(g_title, 1)._style = copy(grid_sty[27][1])
    ws.cell(g_head, 1).value = "Component"
    ws.cell(g_head, 1)._style = copy(grid_sty[28][1])
    ws.cell(g_pct, 1).value = "Consumption %"
    ws.cell(g_pct, 1)._style = copy(grid_sty[29][1])
    for i in range(n):
        c = 2 + i
        ws.cell(g_head, c).value = f"Month {i + 1}"
        ws.cell(g_head, c)._style = copy(grid_sty[28][2])
        ws.cell(g_pct, c).value = f"=B{m_first + i}"
        ws.cell(g_pct, c)._style = copy(grid_sty[29][2])
    ws.cell(g_head, total_col).value = "Total"
    ws.cell(g_head, total_col)._style = copy(grid_sty[28][2])

    last_month_col = get_column_letter(1 + n)
    disc = f"*(1-'Pricing Overview'!{discount_cell})" if discount_cell else ""
    for k, (label, po_rows, discountable) in enumerate(components):
        r = g_first + k
        src = grid_sty[min(30 + k, 36)]
        ws.cell(r, 1).value = label
        ws.cell(r, 1)._style = copy(src[1])
        # Steady-state cost for this line: one or more Pricing Overview baseline cells,
        # shifted for the cloud-bill relayout, discounted only where the discount applies.
        refs = [f"'Pricing Overview'!$B${rr - po_delta}" for rr in po_rows]
        base = f"({'+'.join(refs)})" if len(refs) > 1 else refs[0]
        factor = disc if discountable else ""
        for i in range(n):
            c = 2 + i
            cl = get_column_letter(c)
            ws.cell(r, c).value = f"={base}{factor}*{cl}{g_pct}"
            ws.cell(r, c)._style = copy(src[2])
        ws.cell(r, total_col).value = f"=SUM(B{r}:{last_month_col}{r})"
        ws.cell(r, total_col)._style = copy(src[2])

    ws.cell(g_cum, 1).value = "Cumulative Total"
    ws.cell(g_cum, 1)._style = copy(grid_sty[37][1])
    for i in range(n):
        c = 2 + i
        cl = get_column_letter(c)
        ws.cell(g_cum, c).value = f"=SUM($B${g_first}:{cl}{g_last})"
        ws.cell(g_cum, c)._style = copy(grid_sty[37][2])
    # "Total" for the cumulative row is the year-1 grand total = the last month's cumulative
    # (NOT the sum of every month's running cumulative, which would multiply-count the ramp).
    ws.cell(g_cum, total_col).value = f"={last_month_col}{g_cum}"
    ws.cell(g_cum, total_col)._style = copy(grid_sty[37][2])

    # chart-source table: Month / Cumulative Total, re-pointed at the new cumulative row
    for c in (1, 2):
        ws.cell(c_head, c).value = chart_hdr[0][c - 1]
        ws.cell(c_head, c)._style = copy(chart_hdr[1][c])
    for i in range(n):
        r = c_head + 1 + i
        cl = get_column_letter(2 + i)
        ws.cell(r, 1).value = f"Month {i + 1}"
        ws.cell(r, 1)._style = copy(chart_sty[1])
        ws.cell(r, 2).value = f"={cl}{g_cum}"
        ws.cell(r, 2)._style = copy(chart_sty[2])

    ws["A10"] = f"{n}-Month Consumption Ramp"
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == 10 and merged.max_row == 10:
            ws.unmerge_cells(str(merged))
    ws.merge_cells("A10:D10")
    ws["A11"] = "Month"
    ws["B11"] = "Consumption %"
    ws["C11"] = "Modeled Monthly"
    ws["D11"] = "Cumulative"
    for c in range(5, 10):
        ws.cell(11, c).value = None
        ws.cell(11, c).style = "Normal"
    ws.column_dimensions["A"].width = 28
    for c in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws["F5"] = "Costs use the current Pricing Overview total as the steady-state baseline."
    ws["F6"] = "Consumption percentages are copied from the ramp configured in the app."
    ws["F7"] = "No workload sequencing assumptions are imported from the styling reference."
    ws["A9"] = (
        "Monthly costs below are formula-driven from the current BOM and configured "
        "consumption percentages."
    )
    # Steady-State Monthly / Annual come straight from the Pricing Overview Total Monthly (B22
    # on-prem, B18 in the cloud-bill relayout) and Total Annual (B23 / B19) so the summary ties
    # to the same total the grid ramps.
    ws["B5"] = f"='Pricing Overview'!$B${22 - po_delta}"
    ws["D5"] = f"='Pricing Overview'!$B${23 - po_delta}"
    return n


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def _strip_orphan_drawings(path):
    """Drop package parts nothing can reach, and heal relationships that point at nothing.

    The template ships images and a chart on sheets the Full BOM deletes. Rewriting it leaves
    the abandoned parts and their [Content_Types].xml overrides behind, and Excel treats a part
    it can't reach as unreadable - hence "we found a problem with some content".

    Reachability is computed properly, by walking the relationship graph from the package root
    the way Excel does, instead of pattern-matching one spelling of a target path. An earlier
    version only recognised relative targets ("../drawings/drawing1.xml"); openpyxl writes
    absolute ones ("/xl/drawings/drawing1.xml"), so every drawing looked unreferenced and got
    deleted while the sheets kept pointing at them - which caused the very corruption this is
    meant to prevent. Targets are now resolved in both forms, and anything still dangling has
    both the relationship and the element that referenced it removed.

    openpyxl and LibreOffice tolerate both faults, which is why only Excel ever complained.
    """
    import os as _os, zipfile, re as _re, posixpath, shutil, tempfile as _tf
    try:
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
            payload = {n: z.read(n) for n in z.namelist()}

        def rels_for(part):
            d, b = posixpath.split(part)
            return posixpath.join(d, "_rels", b + ".rels") if part else "_rels/.rels"

        def resolve(base_part, target):
            """A rel target is either absolute ('/xl/...') or relative to its owner's folder."""
            if target.startswith("/"):
                return target.lstrip("/")
            return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))

        rel_re = _re.compile(r"<Relationship\b[^>]*/>|<Relationship\b[^>]*>.*?</Relationship>", _re.S)

        def parse_rels(rels_part):
            """-> [(rel_id, resolved_target_part, raw_xml)] for internal (non-external) rels."""
            out = []
            owner = rels_part.replace("/_rels/", "/", 1)[:-5] if rels_part != "_rels/.rels" else ""
            blob = payload.get(rels_part)
            if not blob:
                return out
            for tag in rel_re.findall(blob.decode("utf-8", "replace")):
                if 'TargetMode="External"' in tag:
                    continue
                tid = _re.search(r'Id="([^"]+)"', tag)
                tgt = _re.search(r'Target="([^"]+)"', tag)
                if not (tid and tgt):
                    continue
                out.append((tid.group(1), resolve(owner, tgt.group(1)), tag))
            return out

        # ---- pass 1: heal dangling relationships (target part is not in the package) -------
        for rels_part in [n for n in list(payload) if n.endswith(".rels")]:
            dangling = [(i, t, tag) for i, t, tag in parse_rels(rels_part)
                        if t not in names]
            if not dangling:
                continue
            xml = payload[rels_part].decode("utf-8", "replace")
            owner = rels_part.replace("/_rels/", "/", 1)[:-5]
            owner_xml = payload.get(owner, b"").decode("utf-8", "replace")
            for rid, _t, tag in dangling:
                xml = xml.replace(tag, "")
                # Remove whatever element pointed at it (<drawing r:id="rId1"/> and friends),
                # otherwise the sheet references a relationship that no longer exists.
                owner_xml = _re.sub(r'<[A-Za-z:]+[^>]*r:id="%s"[^>]*/>' % _re.escape(rid),
                                    "", owner_xml)
            payload[rels_part] = xml.encode("utf-8")
            if owner in payload:
                payload[owner] = owner_xml.encode("utf-8")

        # ---- pass 2: reachability walk from the package root -------------------------------
        reachable = set()
        queue = [""]
        while queue:
            part = queue.pop()
            rp = rels_for(part)
            if rp in payload:
                reachable.add(rp)
            for _rid, target, _tag in parse_rels(rp):
                if target in names and target not in reachable:
                    reachable.add(target)
                    queue.append(target)

        keep_always = {"[Content_Types].xml", "_rels/.rels"}
        drop = {n for n in payload
                if n not in reachable and n not in keep_always and not n.endswith("/")}
        # Only ever prune inside xl/ - docProps and friends are reached by the root rels, and
        # anything unexpected outside xl/ is left alone rather than guessed at.
        drop = {n for n in drop if n.startswith("xl/")}
        for n in drop:
            payload.pop(n, None)

        # Content types must describe exactly what's in the package: drop the overrides for
        # parts we removed, and for any part that was already missing. An override naming a
        # part that isn't there is itself enough to make Excel offer to repair the file.
        ct = payload.get("[Content_Types].xml", b"").decode("utf-8", "replace")
        ct = _re.sub(
            r'<Override[^>]*PartName="/([^"]+)"[^>]*/>',
            lambda m: m.group(0) if m.group(1) in payload else "",
            ct,
        )
        payload["[Content_Types].xml"] = ct.encode("utf-8")

        fd, tmp = _tf.mkstemp(suffix=".xlsx")
        _os.close(fd)
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
            for n, blob in payload.items():
                out.writestr(n, blob)
        shutil.move(tmp, path)
    except Exception:
        # Never let cleanup break a working export.
        return


def build_full_bom_bytes(pricing, rows=None, fields=None, ramp=None, bom_name="",
                         shape=None, hours=None, block_rate=None, vpu_rate=None,
                         default_vpus=None, file_rate=None, windows_rate=None,
                         windows_sku=None, optimization=0.0, include_diagram=True,
                         extra_services=None, cloud_comparison=None, diagram_options=None,
                         workflow_json=None):
    """Build the 12-sheet Full BOM workbook populated from the app's priced inventory.

    The workbook is built to tie out to the app EXACTLY:
      Compute (OCPU+RAM+block) + Storage + 3rd-party licensing = the app's monthly total,
      and the Consumption Ramp follows the app's ramp curve. Sheets the app does not model
      yet (Networking / Security KMS / DR / Backups) are zeroed so they inject no cost.

    EVERY rate is supplied by the caller from the app's catalog. These deliberately have NO
    numeric defaults: a default here is a second copy of the price list, and a stale copy is
    exactly how a deliverable ends up quoting last year's rates. If the caller doesn't pass
    a rate, we pull it from app.py rather than from whatever the source template shipped.
    """
    if None in (hours, block_rate, vpu_rate, default_vpus, file_rate, windows_rate,
                windows_sku):
        import app as _app                       # the single source of truth for pricing
        hours = _app.HOURS_PER_MONTH if hours is None else hours
        block_rate = _app.storage_rate("B91961") if block_rate is None else block_rate
        vpu_rate = _app.storage_rate("B91962") if vpu_rate is None else vpu_rate
        file_rate = _app.storage_rate("B89057") if file_rate is None else file_rate
        default_vpus = (_app.BLOCK_PERFORMANCE_UNITS_PER_GB
                        if default_vpus is None else default_vpus)
        windows_rate = _app.WINDOWS_LICENSE_RATE if windows_rate is None else windows_rate
        windows_sku = _app.WINDOWS_LICENSE_SKU if windows_sku is None else windows_sku
    spec = load_spec()
    wb = build_workbook(spec)
    wb.properties.creator = "OCI BOM + Architecture Generator"
    wb.properties.lastModifiedBy = "OCI BOM + Architecture Generator"
    wb.properties.title = (bom_name or "OCI Bill of Materials").strip()
    wb.properties.subject = "Oracle Cloud Infrastructure bill of materials"
    # These floating notes were commercial/project statements in the reference workbook,
    # not styling. Never carry them into a generated estimate.
    ws_overview = wb["Pricing Overview"]
    for merged in list(ws_overview.merged_cells.ranges):
        if str(merged) == "D23:H25":
            ws_overview.unmerge_cells(str(merged))
    for r in range(23, 26):
        for c in range(4, 9):
            ws_overview.cell(r, c).value = None
            ws_overview.cell(r, c).style = "Normal"
    for r in range(24, 26):
        ws_overview.cell(r, 2).value = None
        ws_overview.cell(r, 2).style = "Normal"

    keys = _resolve_inventory_keys(fields or [])
    raw_by_id = {}
    for r in (rows or []):
        rid = r.get("__id") or r.get("rowId")
        if rid:
            raw_by_id[str(rid)] = r

    servers = []
    apps = []
    storage_rows = []
    # An overview payload is now built for BOTH modes (on-prem uses current on-prem spend as
    # the baseline), so "is this a cloud bill" can no longer be inferred from its presence.
    on_prem_overview = bool((cloud_comparison or {}).get("onPrem"))
    is_cloud_bill = bool(cloud_comparison) and not on_prem_overview
    for pr in (pricing or {}).get("rows", []):
        specs = pr.get("specs") or {}
        raw = raw_by_id.get(str(pr.get("rowId"))) or {}

        # Cloud-bill mode: only the Compute-category bill lines belong on the Compute sheet.
        # Storage/networking/etc. service rows roll into the Pricing Overview lines instead,
        # so including them here would double-count them.
        if is_cloud_bill and (pr.get("ociServiceCategory") or "") != "Compute":
            continue

        def rv(role):
            k = keys.get(role)
            return _clean(raw.get(k)) if k else ""

        virt = rv("virt")
        # Normalize to the template's Virtual/Physical vocabulary (drives the OCPU formula).
        vl = _norm(virt)
        if "phys" in vl:
            virt = "Physical"
        elif vl:
            virt = "Virtual"

        # Column I is "vCPU / Cores" and the sheet's OCPU formula is
        #   =IF(UPPER(TRIM(virt))="PHYSICAL", I, I/2)
        # i.e. a PHYSICAL row's cores are OCPUs 1:1 and only a VIRTUAL row's vCPUs are halved.
        # So write the figure that reproduces the app's OCPU under that formula: physical rows
        # carry their core count (= OCPU), virtual rows carry vCPUs (= OCPU x 2). Writing
        # ocpus*2 unconditionally made the sheet double every physical server's OCPU (and its
        # compute cost) relative to the app.
        _ocpu_val = float(specs.get("ocpus") or 0)
        if virt == "Physical":
            vcpu = _ocpu_val
        else:
            vcpu = specs.get("vcpus") or (_ocpu_val * 2)

        # Only use a REAL application column. If the inventory has none, leave Master
        # Application blank (and the Applications sheet empty) rather than duplicating
        # the server name - no invented data. Cloud bills have no application grouping,
        # so never carry an "app" here (the inventory-key match can otherwise land on an
        # unrelated column like mapping confidence).
        app_name = "" if is_cloud_bill else rv("app")
        srv = {
            "server": rv("server") or _clean(pr.get("name")),
            "tier": rv("tier"),
            "env": rv("env") or _clean(pr.get("environment")),
            "app": app_name,
            "desc": rv("desc"),
            "virt": virt,
            "os_name": rv("os_name"),
            # Prefer a real OS family/type column; otherwise fall back to the same
            # Windows/Linux classification the estimator priced on (osDetected), so the
            # export's OS column matches the Price page and the Windows-license basis.
            "os_family": rv("os_family") or (str(pr.get("osDetected") or "").capitalize() or None),
            # Full precision - rounding here would put the workbook a few cents off the
            # app's total across hundreds of rows, and the two must tie out exactly.
            "vcpu": float(vcpu) if vcpu else None,
            "memory_gb": float(specs.get("memoryGb") or 0) or None,
            "storage_gb": float(specs.get("blockStorageGb") or 0) or None,
            # The OCI shape this server maps to (shown on the Compute sheet). shapeUsed can
            # be a dict (on-prem shape payload) or a string (cloud-bill shape name).
            "shape": _shape_used_label(pr.get("shapeUsed")),
            # The source instance/shape this row maps FROM (shown next to the OCI shape).
            "source_shape": _source_shape_label(pr) if is_cloud_bill else "",
            # Per-row monthly hours from the data source (the app already priced each row at
            # its own hours). Falls back to the global hours only when the row has none.
            # Cloud-bill: use the EFFECTIVE hours implied by the bill's metered usage
            # (OCPU-hours / OCPU) so OCPU x hours x rate reproduces the app's actual cost -
            # a bill line can cover far more than one instance's 730 hours.
            "hours": (_cloud_effective_hours(pr) if is_cloud_bill
                      else float(pr.get("hoursPerMonth") or 0) or None),
        }
        if not (srv["vcpu"] or srv["memory_gb"] or srv["storage_gb"]):
            continue
        servers.append(srv)
        if app_name and app_name not in apps:
            apps.append(app_name)

        fs = float(specs.get("fileStorageGb") or 0)
        if fs > 0:
            storage_rows.append({
                "server": srv["server"], "tier": srv["tier"], "env": srv["env"],
                "app": app_name, "signal": "Shared / file storage",
                "target": "OCI File Storage", "gb": round(fs, 2), "rate": file_rate,
            })

    # Cloud-bill: itemize the storage services on the Storage sheet (the compute-loop above
    # skipped them). This is display only - the Storage line total is set on the Pricing
    # Overview by _add_cloud_bill_services, so there's no double count.
    if is_cloud_bill:
        storage_rows = _cloud_storage_rows(cloud_comparison.get("pricing") or pricing)
    _extra_priced = None
    if extra_services:
        import oci_catalog
        _extra_priced, _ = oci_catalog.price_extras(extra_services, hours)
        storage_rows.extend(_extra_storage_rows(_extra_priced))
    windows_monthly = sum(float(r.get("windowsLicenseMonthly") or 0)
                          for r in (pricing or {}).get("rows", []))
    # Build the Rate Card FIRST - from only the SKUs/rates used in this build, sorted
    # alphabetically - so the Compute/Storage formulas can reference the exact cells it
    # placed each rate on (transparency, and it ties out to the app).
    shape_label = (shape or {}).get("shortLabel") or (shape or {}).get("label") or ""
    rate_entries = _collect_rate_card_entries(
        shape, block_rate, vpu_rate, default_vpus, hours, file_rate, windows_rate,
        windows_sku, windows_monthly > 0, servers, storage_rows, pricing,
        extra_services, is_cloud_bill)
    rate_refs = _write_rate_card(wb["Rate Card"], rate_entries)
    _compute_last = _populate_compute(
        wb[COMPUTE_SHEET], servers, hours, rate_refs, shape_label)
    _populate_apps(wb[APPS_SHEET], apps, servers, shape_label)
    _populate_storage(wb[STORAGE_SHEET], storage_rows, rate_refs, file_rate)
    _set_toc(wb["Table of Contents"], bom_name)
    _set_assumptions(wb["Assumptions"], servers, shape_label, hours)
    _apply_customer_name(wb, bom_name)
    # Compute optimization: record the % the app's Rightsize applied. Ax = 15% OCPU /
    # 20% RAM, regular E6 = 10% / 15%; the app already shrank the quantities to match.
    rightsized = any(r.get("rightsized") for r in (pricing or {}).get("rows", []))
    shape_key = str((shape or {}).get("key") or "")
    is_ax = shape_key.endswith("-ax")
    if is_ax:
        ocpu_pct, ram_pct = 0.15, 0.20
    elif shape_key == "e6-standard":
        ocpu_pct, ram_pct = 0.10, 0.15
    else:
        ocpu_pct = ram_pct = 0.0
    _set_optimization(wb[COMPUTE_SHEET], rightsized, ocpu_pct, ram_pct, is_ax)
    _zero_unmodeled_sheets(wb)
    _add_licensing_line(wb["Pricing Overview"], windows_monthly)
    # App-added OCI services roll into the matching Pricing Overview lines (which sit inside
    # the total) and are itemized on the Networking sheet.
    if _extra_priced:
        _add_extra_services(wb, _extra_priced)
    # Cloud-bill mode: roll the non-compute mapped services into the Pricing Overview lines
    # so the template total covers the whole bill, not just compute.
    if is_cloud_bill:
        _add_cloud_bill_services(wb, cloud_comparison.get("pricing") or pricing)
    _service_pricing = (
        cloud_comparison.get("pricing") if is_cloud_bill else pricing) or pricing
    _write_service_tabs(
        wb, _service_pricing, _extra_priced, spec, _compute_last, len(storage_rows))
    # Cloud-bill mode shifts the Pricing Overview baseline up by 4 rows (relayout below) and
    # exposes an editable OCI discount at $B$7; feed both to the ramp so its lines point at the
    # right cells and tie to Total Monthly. On-prem keeps the template layout (no shift/discount).
    _ramp_delta = 4 if cloud_comparison else 0
    _ramp_discount = "$B$7" if cloud_comparison else None
    ramp_months = _populate_ramp(wb["Consumption Ramp"], ramp, include_windows=windows_monthly > 0,
                                 po_delta=_ramp_delta, discount_cell=_ramp_discount)
    _repoint_ramp_refs(wb["Pricing Overview"], ramp_months, include_windows=windows_monthly > 0)

    # Re-layout the Pricing Overview ONLY in cloud-bill mode, where the AWS-vs-OCI
    # comparison blocks need to stack beneath the baseline. Shift the Baseline + OCI Cost
    # Profile band up by 4 (rows 12-23 -> 8-19, closing the gap); Total Monthly -> B18,
    # Total Annual -> B19. On-prem BOMs have no comparison blocks, so the template's
    # original layout (baseline at 12-23, diagram at row 29) is left untouched.
    _relayout_notes = None
    if cloud_comparison:
        _relayout_notes = _relayout_pricing_overview(wb["Pricing Overview"], delta=4)
        # OCI discount input (A7 label / B7 editable %). The Total Monthly applies it LIVE to
        # the discountable OCI services and leaves 3rd-party licensing (Windows + SQL Server) at
        # list - matching the app's headline math. Everything downstream (Total Annual, the
        # comparison blocks, chart) references B18/B19, so editing B7 re-flows the whole sheet.
        # SQL licensing is bundled in service categories, so split it by group (Compute -> B11,
        # Database/other -> B12) so the discount can exclude it exactly.
        import bom_export as _bx_disc
        _sql_compute = _sql_database = 0.0
        for _r in ((cloud_comparison.get("pricing") or pricing).get("rows") or []):
            if (_r.get("costAction") or "") == "remove":
                continue
            _s = float(_r.get("sqlLicenseMonthly") or 0)
            if not _s:
                continue
            _g = _bx_disc._cloud_product_group(_r.get("ociServiceCategory"), _r.get("sourceService"))
            if _g == "Compute":
                _sql_compute += _s
            else:
                _sql_database += _s
        _apply_overview_discount(wb["Pricing Overview"], float(cloud_comparison.get("ociDiscount") or 0),
                                 _sql_compute, _sql_database)

    # Cloud-bill: add the AWS-vs-OCI comparison blocks (5-year projection, savings,
    # chart) directly below the baseline on the Pricing Overview, wired to its live
    # B18/B19 totals (post-relayout).
    if cloud_comparison:
        import bom_export
        _cc_pricing = cloud_comparison.get("pricing") or pricing
        # On-prem compares against what the estate costs to run today; cloud-bill against the
        # uploaded bill. Either way the figure lands in an editable cell.
        if on_prem_overview:
            _baseline_monthly = float(cloud_comparison.get("baselineMonthly") or 0)
            _baseline_cloud, _baseline_estimated = "onprem", False
        else:
            _baseline_monthly = float((_cc_pricing.get("totals") or {}).get("sourceMonthlyCost") or 0)
            _baseline_cloud = _cc_pricing.get("sourceCloud") or "aws"
            _baseline_estimated = bool(_cc_pricing.get("sourceCostEstimated"))
        bom_export.add_comparison_to_pricing_overview(
            wb["Pricing Overview"], 22, "$B$18", "$B$19",
            _baseline_monthly, bom_export._util_by_year(cloud_comparison.get("ramp")),
            source_cloud=_baseline_cloud, estimated=_baseline_estimated)

    # Architecture diagram generated from THIS BOM. The optional AI plan is constrained
    # metadata; graph geometry, quantities, rendering, and validation remain deterministic.
    # If the diagram toolchain isn't available the export still succeeds - the template's
    # reference picture just stays in place.
    arch_png = None
    if include_diagram:
        try:
            import bom_diagram
            arch_drawio, arch_png = bom_diagram.build_architecture(
                pricing, rows, keys, bom_name,
                (shape or {}).get("shortLabel") or (shape or {}).get("label") or "",
                sites=_distinct_sites(fields or [], rows or []),
                extra_priced=_extra_priced, diagram_options=diagram_options or {})
            if arch_png:
                import app as _app_qa
                _architecture_qa = _app_qa.architecture_artifact_qa(
                    arch_drawio, arch_png
                )
                if not _architecture_qa["passed"]:
                    raise ValueError(
                        "Architecture output failed validation: "
                        + " ".join(_architecture_qa["issues"])
                    )
                # Drop the diagram to the BOTTOM of the sheet, below the comparison blocks.
                # The spec anchor sits at Excel rows 29-72; shift it down so it starts at
                # ~row 52 (2 rows below the comparison area) and keeps its 43-row height.
                import copy as _copy
                arch_anchor = spec.get("architecture_anchor")
                if arch_anchor and cloud_comparison:
                    # Cloud-bill only: the baseline was shifted up and the comparison
                    # blocks now occupy ~rows 22-50, so drop the diagram below them to
                    # ~row 52 (keeping its 43-row height). On-prem keeps the spec anchor
                    # (row 29) since nothing moved.
                    arch_anchor = _copy.deepcopy(arch_anchor)
                    off = 58 - 29          # drop below the comparison block (now ends ~row 55)
                    arch_anchor["from"]["row"] += off
                    if "to" in arch_anchor:
                        arch_anchor["to"]["row"] += off
                _diag_bottom = embed_architecture(wb["Pricing Overview"], arch_png, arch_anchor)
                # Drop the re-homed footnotes just below the diagram (its height varies with
                # the DR section, so this can't be a fixed row).
                if cloud_comparison and _relayout_notes and _diag_bottom:
                    _place_overview_notes(wb["Pricing Overview"], _diag_bottom + 1, _relayout_notes)
        except Exception:
            # Don't swallow the reason silently - a missing diagram was undebuggable.
            import traceback
            traceback.print_exc()
            arch_png = None
    if not arch_png:
        # Nothing to show: the template ships no architecture picture of its own, so the
        # slot is simply empty. Clear its caption too - at whichever row it now lives on.
        # _relayout_pricing_overview re-homes the caption from D27 to D57, and the comparison
        # block's 5-Year Projection then OWNS D27 (the Year-4 cumulative). Blanking row 27
        # unconditionally wiped that cell and broke the cumulative chain, so only clear the
        # row the caption actually occupies.
        ws_po = wb["Pricing Overview"]
        _cap_row = 57 if _relayout_notes is not None else 27
        for merged in list(ws_po.merged_cells.ranges):
            if str(merged) in (f"D{_cap_row}:P{_cap_row}", f"D{_cap_row}:Q{_cap_row}"):
                ws_po.unmerge_cells(str(merged))
        for c in range(4, 17):
            ws_po.cell(_cap_row, c).value = None
            ws_po.cell(_cap_row, c).style = "Normal"
        # No diagram to anchor to - still place the re-homed footnotes just below the
        # comparison blocks so they aren't lost.
        if cloud_comparison and _relayout_notes:
            _place_overview_notes(ws_po, 52, _relayout_notes)

    # Cloud-bill mode: append the Service Mapping (per-line breakdown) + Notes sheets alongside
    # the 12-sheet deliverable. The Pricing Overview then PULLS its comparison source/OCI/savings
    # straight from the Service Mapping total row, so the two always tie out.
    # On-prem has no uploaded bill to map, so these sheets don't exist and the comparison keeps
    # the editable baseline cell written by add_comparison_to_pricing_overview - repointing it
    # at a 'Service Mapping' sheet that was never created would leave B41 a broken reference.
    if is_cloud_bill:
        try:
            import bom_export
            _cc = bom_export.add_cloud_comparison_sheets(
                wb, cloud_comparison.get("pricing") or {"rows": rows, "totals": {}},
                cloud_comparison.get("ramp"), bom_name,
                cloud_comparison.get("ociDiscount") or 0.0,
                cloud_comparison.get("extraServices"),
                cloud_comparison.get("hours") or hours, use_active=False)
            # Point the Pricing Overview comparison at the Service Mapping totals (single source
            # of truth). sv = comparison start(22)+18 -> B41 = source spend, B42 = OCI.
            _smrow = (_cc or {}).get("serviceMappingTotalRow")
            if _smrow:
                _po = wb["Pricing Overview"]
                _po["B41"] = f"='Service Mapping'!$F${_smrow}"   # Current source spend
                _po["B42"] = f"='Service Mapping'!$H${_smrow}"   # OCI (net, incl. Windows licensing)
                # B43 savings already = B41-B42, so it follows automatically.
        except Exception:
            import traceback
            traceback.print_exc()
        # Service Mapping reads right after the Pricing Overview.
        _place_after(wb, "Service Mapping", "Pricing Overview")

    # Hide any printout sheet that ended up with no data (empty Storage/Networking/DR/
    # Security KMS/Applications/Annexure), so the deliverable doesn't ship blank sections.
    _hide_empty_sheets(wb, apps, storage_rows)

    # Rebuild the Table of Contents to list exactly the sheets this workbook ended up with
    # (the template shipped a stale, misaligned list). Runs after hides so it skips blanks.
    try:
        _rebuild_toc(wb, bom_name)
    except Exception:
        import traceback
        traceback.print_exc()

    # Embed the app workflow (hidden _workflow sheet) so this Full BOM can be re-imported
    # via "Load previous BOM" - same as the Quick/comparison export.
    if workflow_json:
        try:
            import bom_export
            bom_export.embed_workflow_state(wb, workflow_json)
        except Exception:
            import traceback
            traceback.print_exc()

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tmp_path = tf.name
    wb.save(tmp_path)
    _postprocess(tmp_path, spec)
    _strip_orphan_drawings(tmp_path)
    data = Path(tmp_path).read_bytes()
    Path(tmp_path).unlink(missing_ok=True)
    return data
